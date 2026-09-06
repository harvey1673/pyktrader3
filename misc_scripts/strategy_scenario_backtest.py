"""Before/after backtests for strategy signal-weight scenarios.

The module keeps production strategy JSON files read-only. Strategy structure
and non-weight signal settings come from JSON, while both current and proposed
weights come from the ``signal_weights`` workbook created by
:mod:`misc_scripts.signal_weight_excel`. The strategy-level scaler comes from the
JSON file's ``config.pos_scaler`` value.

The backtest core is deliberately independent of the production database.  It
works with any ``SignalProvider`` that returns unit-weight holdings and gross
asset PNL.  ``FactorFrameSignalProvider`` is the default implementation for
historical factor matrices. ``build_generated_historical_provider`` is the
default adapter and constructs signals from saved futures, historical
``spot_df``, routing registries, and ``signal_store`` recipes.
Signal holdings and execution-adjusted PNL reuse ``btmetrics.MetricsBase`` and
the matching ``signal_execution_config`` entry used by the production notebook.
"""

from __future__ import annotations

import argparse
import copy
import datetime as dt
import html as html_lib
import json
import math
import re
import sys
from dataclasses import dataclass, field, replace
from pathlib import Path
from typing import Any, Callable, Dict, Iterable, Mapping, Protocol, Sequence

if __package__ in {None, ""}:
    # Make repository packages importable when this file is invoked directly
    # as ``python misc_scripts/strategy_scenario_backtest.py``.
    project_root = str(Path(__file__).resolve().parents[1])
    if project_root not in sys.path:
        sys.path.insert(0, project_root)

import numpy as np
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_SHEET_NAME = "signal_weights"
BUSINESS_DAYS_PER_YEAR = 244
DEFAULT_COST_RATE = 2e-4
BOND_COST_RATE = 0.6e-4
DEFAULT_PNL_TENORS = ("3m", "6m", "1y", "2y", "3y", "4y", "5y", "7y", "9y", "10y")


@dataclass(frozen=True)
class SignalSpec:
    """One factor entry from ``config.factor_repo``."""

    factor_name: str
    name: str
    type: str
    weight: float
    exec_assets: tuple[str, ...] = ()
    threshold: float = 0.0
    rebal: int | str = 1
    param: tuple[float, ...] = (0.0, 0.0)

    def cache_key(self) -> tuple[Any, ...]:
        """Return the weight-independent identity used for signal caching."""

        return (
            self.factor_name,
            self.name,
            self.type,
            self.exec_assets,
            self.threshold,
            str(self.rebal),
            self.param,
        )


@dataclass(frozen=True)
class StrategyScenario:
    """A complete in-memory strategy definition."""

    name: str
    strategy_file: str
    scaler: float
    signals: Mapping[str, SignalSpec]
    config: Mapping[str, Any] = field(default_factory=dict)
    source_path: Path | None = None


@dataclass(frozen=True)
class SignalBacktestResult:
    """Unit-weight signal paths returned by a :class:`SignalProvider`."""

    holdings: pd.DataFrame
    gross_asset_pnl: pd.DataFrame
    cost_rates: pd.Series
    execution_bucket: str = "default"


class SignalProvider(Protocol):
    """Interface used by the scenario composer."""

    def backtest(self, spec: SignalSpec) -> SignalBacktestResult:
        ...


@dataclass(frozen=True)
class PortfolioResult:
    scenario: StrategyScenario
    cost_mode: str
    signal_source: str
    gross_asset_pnl: pd.DataFrame
    costs_by_asset: pd.DataFrame
    net_asset_pnl: pd.DataFrame
    aggregate_holdings: pd.DataFrame
    trade_volume: pd.DataFrame
    gross_exposure: pd.DataFrame
    signal_pnl: pd.DataFrame
    signal_asset_pnl: Mapping[str, pd.DataFrame]

    @property
    def portfolio_pnl(self) -> pd.Series:
        return self.net_asset_pnl.sum(axis=1).rename(self.scenario.name)


@dataclass(frozen=True)
class BtMetricsResult:
    """Notebook-compatible portfolio and asset statistics for one scenario."""

    portfolio: pd.DataFrame
    assets: pd.DataFrame
    tenors: tuple[str, ...]
    engine: str


@dataclass(frozen=True)
class ScenarioComparison:
    baseline: PortfolioResult
    proposed: PortfolioResult
    summary: pd.DataFrame
    asset_comparison: pd.DataFrame
    signal_attribution: pd.DataFrame
    daily_pnl: pd.DataFrame
    checks: pd.DataFrame
    current_btmetrics: BtMetricsResult
    proposed_btmetrics: BtMetricsResult


def _finite_float(value: Any, label: str) -> float:
    if isinstance(value, bool):
        raise ValueError(f"{label} must be numeric")
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{label} must be numeric") from exc
    if not math.isfinite(result):
        raise ValueError(f"{label} must be finite")
    return result


def _parse_signal_spec(factor_name: str, value: Mapping[str, Any]) -> SignalSpec:
    missing = [key for key in ("name", "type", "weight") if key not in value]
    if missing:
        raise ValueError(
            f"Factor '{factor_name}' is missing required fields: {', '.join(missing)}"
        )
    exec_assets = value.get("exec_assets", [])
    param = value.get("param", [0.0, 0.0])
    if not isinstance(exec_assets, list):
        raise ValueError(f"Factor '{factor_name}': exec_assets must be a list")
    if not isinstance(param, list):
        raise ValueError(f"Factor '{factor_name}': param must be a list")
    return SignalSpec(
        factor_name=str(factor_name),
        name=str(value["name"]).strip(),
        type=str(value["type"]).strip(),
        weight=_finite_float(value["weight"], f"Factor '{factor_name}' weight"),
        exec_assets=tuple(str(asset) for asset in exec_assets),
        threshold=_finite_float(
            value.get("threshold", 0.0), f"Factor '{factor_name}' threshold"
        ),
        rebal=value.get("rebal", 1),
        param=tuple(_finite_float(item, f"Factor '{factor_name}' param") for item in param),
    )


def load_strategy_scenario(
    settings_dir: str | Path,
    strategy_file: str,
    *,
    scaler: float | None = None,
    scenario_name: str = "current",
) -> StrategyScenario:
    """Load one strategy JSON without modifying it."""

    settings_dir = Path(settings_dir)
    if Path(strategy_file).name != strategy_file:
        raise ValueError("strategy_file must be a file name, not a path")
    if Path(strategy_file).suffix.lower() != ".json":
        strategy_file += ".json"
    strategy_path = settings_dir / strategy_file
    if not strategy_path.is_file():
        raise FileNotFoundError(f"Strategy JSON does not exist: {strategy_path}")

    data = json.loads(strategy_path.read_text(encoding="utf-8"))
    config = data.get("config")
    if not isinstance(config, dict):
        raise ValueError(f"Missing object 'config' in {strategy_path}")
    repo = config.get("factor_repo")
    if not isinstance(repo, dict):
        raise ValueError(f"Missing object 'config.factor_repo' in {strategy_path}")

    signals: Dict[str, SignalSpec] = {}
    for factor_name, value in repo.items():
        if not isinstance(value, dict):
            raise ValueError(
                f"Factor '{factor_name}' must contain an object in {strategy_path}"
            )
        signals[str(factor_name)] = _parse_signal_spec(str(factor_name), value)

    selected_scaler = config.get("pos_scaler", 1.0) if scaler is None else scaler
    return StrategyScenario(
        name=scenario_name,
        strategy_file=strategy_file,
        scaler=_finite_float(selected_scaler, "strategy scaler"),
        signals=signals,
        config=copy.deepcopy(config),
        source_path=strategy_path.resolve(),
    )


def _normalized_headers(worksheet: Any) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for column, cell in enumerate(worksheet[1], start=1):
        if cell.value is None:
            continue
        key = str(cell.value).strip().lower().replace(" ", "_")
        if key in headers:
            raise ValueError(f"Duplicate Excel column: {key}")
        headers[key] = column
    required = (
        "strategy",
        "factor_name",
        "signal_name",
        "type",
        "curr_weight",
        "new_weight",
    )
    missing = [key for key in required if key not in headers]
    if missing:
        raise ValueError("Excel sheet is missing columns: " + ", ".join(missing))
    return headers


def load_excel_weight_scenarios(
    template: StrategyScenario,
    excel_path: str | Path,
    *,
    sheet_name: str = DEFAULT_SHEET_NAME,
    current_name: str = "current",
    proposed_name: str = "proposed",
) -> tuple[StrategyScenario, StrategyScenario]:
    """Build current and proposed scenarios from matching workbook rows.

    JSON weights are deliberately ignored. Existing JSON factors provide only
    their non-workbook settings: ``exec_assets``, ``threshold``, ``rebal`` and
    ``param``. New factors receive the standard defaults. A workbook row's
    ``curr_weight`` and ``new_weight`` independently control inclusion in the
    two scenarios; set either value to zero to exclude that sleeve.
    """

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Excel sheet '{sheet_name}' not found")
        worksheet = workbook[sheet_name]
        headers = _normalized_headers(worksheet)
        current_signals: Dict[str, SignalSpec] = {}
        proposed_signals: Dict[str, SignalSpec] = {}
        seen: set[str] = set()
        matched_rows = 0

        accepted_names = {
            template.strategy_file.lower(),
            Path(template.strategy_file).stem.lower(),
        }
        for row_number in range(2, worksheet.max_row + 1):
            strategy_value = worksheet.cell(row_number, headers["strategy"]).value
            if strategy_value is None:
                continue
            strategy_name = str(strategy_value).strip().lower()
            if strategy_name not in accepted_names:
                continue
            matched_rows += 1
            factor_value = worksheet.cell(row_number, headers["factor_name"]).value
            signal_value = worksheet.cell(row_number, headers["signal_name"]).value
            type_value = worksheet.cell(row_number, headers["type"]).value
            current_weight_value = worksheet.cell(
                row_number, headers["curr_weight"]
            ).value
            proposed_weight_value = worksheet.cell(
                row_number, headers["new_weight"]
            ).value
            if factor_value is None or not str(factor_value).strip():
                raise ValueError(f"Row {row_number}: factor_name must not be blank")
            factor_name = str(factor_value).strip()
            if factor_name in seen:
                raise ValueError(
                    f"Row {row_number}: duplicate factor '{factor_name}' for "
                    f"{template.strategy_file}"
                )
            seen.add(factor_name)
            if signal_value is None or not str(signal_value).strip():
                raise ValueError(f"Row {row_number}: signal_name must not be blank")
            if type_value is None or not str(type_value).strip():
                raise ValueError(f"Row {row_number}: type must not be blank")

            current_weight = _finite_float(
                current_weight_value, f"Row {row_number} curr_weight"
            )
            proposed_weight = _finite_float(
                proposed_weight_value, f"Row {row_number} new_weight"
            )
            if factor_name in template.signals:
                source = template.signals[factor_name]
                current_spec = replace(
                    source,
                    name=str(signal_value).strip(),
                    type=str(type_value).strip(),
                    weight=current_weight,
                )
            else:
                current_spec = SignalSpec(
                    factor_name=factor_name,
                    name=str(signal_value).strip(),
                    type=str(type_value).strip(),
                    weight=current_weight,
                )
            current_signals[factor_name] = current_spec
            proposed_signals[factor_name] = replace(
                current_spec, weight=proposed_weight
            )
        if matched_rows == 0:
            raise ValueError(
                f"No rows found for strategy '{template.strategy_file}' in {excel_path}"
            )
    finally:
        workbook.close()

    return (
        replace(template, name=current_name, signals=current_signals),
        replace(template, name=proposed_name, signals=proposed_signals),
    )


def apply_excel_proposal(
    baseline: StrategyScenario,
    excel_path: str | Path,
    *,
    sheet_name: str = DEFAULT_SHEET_NAME,
    scenario_name: str = "proposed",
) -> StrategyScenario:
    """Compatibility wrapper returning the workbook's proposed scenario."""

    _, proposed = load_excel_weight_scenarios(
        baseline,
        excel_path,
        sheet_name=sheet_name,
        proposed_name=scenario_name,
    )
    return proposed


def _xs_demean(frame: pd.DataFrame) -> pd.DataFrame:
    return frame.sub(frame.mean(axis=1), axis=0)


def _xs_score(frame: pd.DataFrame) -> pd.DataFrame:
    demeaned = _xs_demean(frame)
    scale = demeaned.std(axis=1).replace(0.0, np.nan)
    return demeaned.div(scale, axis=0)


def _xs_rank(frame: pd.DataFrame, cutoff: float) -> pd.DataFrame:
    counts = frame.count(axis=1).replace(0, np.nan)
    ranks = frame.rank(axis=1)
    median = ranks.quantile(0.5, axis=1)
    result = ranks.sub(median, axis=0).div(counts, axis=0) * 2.0
    if cutoff < 0.5:
        result = result.where(result.abs() > (1.0 - cutoff * 2.0), 0.0)
    return result


def _apply_rebalance(frame: pd.DataFrame, rebal: int | str) -> pd.DataFrame:
    if isinstance(rebal, str):
        value = rebal.strip().lower()
        if len(value) <= 3 or not value[3:].isdigit():
            raise ValueError(f"Unsupported rebal value: {rebal}")
        window = int(value[3:])
        method = value[:3]
        if method == "sma":
            return frame.rolling(window).mean()
        if method == "ema":
            return frame.ewm(window).mean()
        raise ValueError(f"Unsupported rebal value: {rebal}")
    window = int(rebal)
    if window < 1:
        raise ValueError("rebal must be at least 1")
    return frame.rolling(window).mean()


def _execution_signal_name(spec: SignalSpec) -> str:
    """Return the signal-repo key used by ``signal_execution_config``."""

    signal_type = spec.type.strip().lower()
    if "xs" not in signal_type:
        return spec.name
    suffix = signal_type.split("-", 1)[1] if "-" in signal_type else "rank"
    suffix_map = {
        "demean": "xdemean",
        "xdemean": "xdemean",
        "demedian": "xdemedian",
        "score": "xscore",
        "xscore": "xscore",
        "rank": "xrank",
        "rank_cutoff": "xrank",
    }
    return f"{spec.name}_{suffix_map.get(suffix, suffix)}"


class FactorFrameSignalProvider:
    """Backtest factors supplied as historical date-by-asset matrices.

    ``exec_assets`` follows current production behavior and is treated as an
    exclusion list. Holdings and PNL are calculated by ``btmetrics.MetricsBase``.
    ``holding_lag=2`` reproduces the notebook's two-stage lag: one signal shift
    and one execution lag.  A metrics class can be injected for dependency-light
    unit tests; production lazily imports the repository implementation.
    """

    def __init__(
        self,
        factor_values: Mapping[str, pd.DataFrame],
        asset_returns: pd.DataFrame,
        volatility: pd.DataFrame,
        cost_rates: Mapping[str, float] | pd.Series,
        *,
        holding_lag: int = 2,
        execution_buckets: Mapping[str, str] | None = None,
        close_prices: pd.DataFrame | None = None,
        execution_prices: Mapping[str, pd.DataFrame] | None = None,
        execution_config: Mapping[str, Mapping[str, Any]] | None = None,
        volatility_overrides: Mapping[str, pd.DataFrame] | None = None,
        traded_price_overrides: Mapping[str, pd.DataFrame] | None = None,
        pnl_modes: Mapping[str, str] | None = None,
        post_funcs: Mapping[str, str] | None = None,
        default_execution_window: str = "close",
        metrics_class: type | None = None,
        signal_buffer_func: Callable[[pd.DataFrame, float], pd.DataFrame] | None = None,
        signal_cost_optim_func: Callable[..., pd.DataFrame] | None = None,
    ) -> None:
        self.factor_values = dict(factor_values)
        self.asset_returns = asset_returns.sort_index().copy()
        self.volatility = volatility.sort_index().copy()
        self.cost_rates = pd.Series(cost_rates, dtype=float)
        self.holding_lag = int(holding_lag)
        self.execution_buckets = dict(execution_buckets or {})
        self.close_prices = None if close_prices is None else close_prices.sort_index().copy()
        self.execution_prices = {
            name: frame.sort_index().copy()
            for name, frame in (execution_prices or {}).items()
        }
        self.execution_config = {
            name: dict(value) for name, value in (execution_config or {}).items()
        }
        self.volatility_overrides = {
            name: frame.sort_index().copy()
            for name, frame in (volatility_overrides or {}).items()
        }
        self.traded_price_overrides = {
            name: frame.sort_index().copy()
            for name, frame in (traded_price_overrides or {}).items()
        }
        self.pnl_modes = dict(pnl_modes or {})
        self.post_funcs = dict(post_funcs or {})
        self.default_execution_window = default_execution_window
        self.metrics_class = metrics_class
        self.signal_buffer_func = signal_buffer_func
        self.signal_cost_optim_func = signal_cost_optim_func
        self._cache: Dict[tuple[Any, ...], SignalBacktestResult] = {}

    def _metrics_type(self) -> type:
        if self.metrics_class is not None:
            return self.metrics_class
        try:
            from pycmqlib3.analytics.btmetrics import MetricsBase
        except ImportError as exc:
            raise ImportError(
                "Unable to import pycmqlib3.analytics.btmetrics.MetricsBase. "
                "Run the scenario backtest in the production Python environment "
                "with the repository analytics dependencies installed."
            ) from exc
        self.metrics_class = MetricsBase
        return MetricsBase

    def backtest(self, spec: SignalSpec) -> SignalBacktestResult:
        key = spec.cache_key()
        if key in self._cache:
            return self._cache[key]
        factor_key = (
            spec.factor_name
            if spec.factor_name in self.factor_values
            else spec.name
        )
        if factor_key not in self.factor_values:
            raise KeyError(
                f"Historical factor data is unavailable for signal '{spec.name}' "
                f"(factor '{spec.factor_name}')"
            )

        signal = self.factor_values[factor_key].sort_index().copy()
        signal.index = pd.to_datetime(signal.index)
        signal_volatility = self.volatility_overrides.get(factor_key, self.volatility)
        common_assets = signal.columns.intersection(self.asset_returns.columns)
        common_assets = common_assets.intersection(signal_volatility.columns)
        common_assets = common_assets.difference(list(spec.exec_assets))
        if len(common_assets) == 0:
            raise ValueError(
                f"Signal '{spec.factor_name}' has no tradable assets after exclusions"
            )
        signal = signal.loc[:, common_assets]

        signal_type = spec.type.lower()
        if signal_type != "pos":
            if "xs" in signal_type:
                suffix = signal_type.split("-", 1)[1] if "-" in signal_type else "rank"
                if suffix in {"demean", "xdemean"}:
                    signal = _xs_demean(signal)
                elif suffix == "demedian":
                    signal = signal.sub(signal.median(axis=1), axis=0)
                elif suffix == "xscore":
                    signal = _xs_score(signal)
                elif suffix in {"rank", "rank_cutoff"}:
                    signal = _xs_rank(signal, spec.threshold)
                else:
                    raise ValueError(
                        f"Unsupported cross-sectional type '{spec.type}' for "
                        f"factor '{spec.factor_name}'"
                    )
            elif signal_type != "ts":
                raise ValueError(
                    f"Unsupported signal type '{spec.type}' for factor "
                    f"'{spec.factor_name}'"
                )
            signal = _apply_rebalance(signal, spec.rebal)

        execution_name = _execution_signal_name(spec)
        post_func = self.post_funcs.get(execution_name, "")
        last_post_func = post_func.split("|")[-1].strip().lower()
        if last_post_func.startswith("buf"):
            if self.signal_buffer_func is None:
                from pycmqlib3.analytics.tstool import signal_buffer

                self.signal_buffer_func = signal_buffer
            signal = self.signal_buffer_func(signal, float(last_post_func[3:]))
        elif last_post_func.startswith("bfc"):
            if self.signal_cost_optim_func is None:
                from pycmqlib3.analytics.tstool import signal_cost_optim

                self.signal_cost_optim_func = signal_cost_optim
            post_volatility = signal_volatility.reindex(
                index=signal.index, columns=common_assets
            ).ffill()
            signal = self.signal_cost_optim_func(
                signal,
                float(last_post_func[3:]),
                post_volatility,
                cost_dict=self.cost_rates.reindex(common_assets).fillna(DEFAULT_COST_RATE).to_dict(),
                turnover_dict={},
                power=3,
            )

        returns = self.asset_returns.loc[:, common_assets]
        volatility = signal_volatility.loc[:, common_assets].replace(0.0, np.nan)
        index = signal.index.intersection(returns.index).intersection(volatility.index)
        if len(index) == 0:
            raise ValueError(f"Signal '{spec.factor_name}' has no overlapping dates")
        signal = signal.reindex(index=index).ffill().fillna(0.0)
        returns = returns.reindex(index=index).fillna(0.0)
        volatility = volatility.reindex(index=index).ffill()
        base_holdings = signal.div(volatility).replace([np.inf, -np.inf], np.nan)
        base_holdings = base_holdings.shift(1)
        execution_settings = self.execution_config.get(
            execution_name,
            {
                "win": self.default_execution_window,
                "lag": max(0, self.holding_lag - 1),
            },
        )
        execution_window = str(execution_settings.get("win", self.default_execution_window))
        execution_lag = int(execution_settings.get("lag", max(0, self.holding_lag - 1)))
        if execution_lag < 0:
            raise ValueError(f"Execution lag must not be negative for '{execution_name}'")

        metrics_type = self._metrics_type()
        zero_costs = {asset: 0.0 for asset in common_assets}
        if self.close_prices is not None:
            traded_override = self.traded_price_overrides.get(factor_key)
            if traded_override is None and execution_window not in self.execution_prices:
                raise KeyError(
                    f"Execution-price window '{execution_window}' is unavailable "
                    f"for signal '{execution_name}'"
                )
            close_prices = self.close_prices.reindex(index=index, columns=common_assets).ffill()
            traded_source = (
                traded_override
                if traded_override is not None
                else self.execution_prices[execution_window]
            )
            traded_prices = traded_source.reindex(index=index, columns=common_assets).ffill()
            pnl_mode = self.pnl_modes.get(factor_key, "ret")
            if pnl_mode not in {"ret", "px"}:
                raise ValueError(f"Unsupported PNL mode '{pnl_mode}' for '{spec.name}'")
            metric_returns = traded_prices.diff() if pnl_mode == "px" else traded_prices.pct_change()
            metrics = metrics_type(
                holdings=base_holdings,
                returns=metric_returns,
                shift_holdings=execution_lag,
                cost_dict=zero_costs,
            )
            holdings = metrics.holdings.fillna(0.0)
            if pnl_mode == "px":
                gross_asset_pnl = metrics.calculate_pnl_stats(
                    shift=0, tenors=False, perf_metrics=[]
                )["asset_pnl"].reindex_like(holdings).fillna(0.0)
            else:
                gross_asset_pnl = metrics.calculate_daily_pnl(
                    traded_prices, close_prices, mode="ret"
                ).reindex_like(holdings).fillna(0.0)
        else:
            metrics = metrics_type(
                holdings=base_holdings,
                returns=returns,
                shift_holdings=execution_lag,
                cost_dict=zero_costs,
            )
            holdings = metrics.holdings.fillna(0.0)
            gross_asset_pnl = metrics.calculate_pnl_stats(
                shift=0, tenors=False, perf_metrics=[]
            )["asset_pnl"].reindex_like(holdings).fillna(0.0)
        rates = self.cost_rates.reindex(common_assets).fillna(DEFAULT_COST_RATE)
        result = SignalBacktestResult(
            holdings=holdings,
            gross_asset_pnl=gross_asset_pnl,
            cost_rates=rates,
            execution_bucket=self.execution_buckets.get(
                spec.factor_name, execution_window
            ),
        )
        self._cache[key] = result
        return result


def _validated_path(result: SignalBacktestResult, factor_name: str) -> SignalBacktestResult:
    if result.holdings.empty or result.gross_asset_pnl.empty:
        raise ValueError(f"Signal '{factor_name}' returned an empty backtest")
    holdings = result.holdings.copy().sort_index()
    pnl = result.gross_asset_pnl.copy().sort_index()
    holdings.index = pd.to_datetime(holdings.index)
    pnl.index = pd.to_datetime(pnl.index)
    index = holdings.index.intersection(pnl.index)
    columns = holdings.columns.intersection(pnl.columns)
    if len(index) == 0 or len(columns) == 0:
        raise ValueError(f"Signal '{factor_name}' holdings and PNL do not overlap")
    holdings = holdings.loc[index, columns].apply(pd.to_numeric, errors="coerce")
    pnl = pnl.loc[index, columns].apply(pd.to_numeric, errors="coerce")
    if np.isinf(holdings.to_numpy()).any() or np.isinf(pnl.to_numpy()).any():
        raise ValueError(f"Signal '{factor_name}' returned infinite values")
    return SignalBacktestResult(
        holdings=holdings.fillna(0.0),
        gross_asset_pnl=pnl.fillna(0.0),
        cost_rates=result.cost_rates.reindex(columns).fillna(DEFAULT_COST_RATE),
        execution_bucket=result.execution_bucket,
    )


def compose_portfolio(
    scenario: StrategyScenario,
    provider: SignalProvider,
    *,
    cost_mode: str = "netted",
    cost_multiplier: float = 1.0,
) -> PortfolioResult:
    """Backtest and aggregate a complete strategy scenario."""

    if cost_mode not in {"netted", "sleeve"}:
        raise ValueError("cost_mode must be 'netted' or 'sleeve'")
    cost_multiplier = _finite_float(cost_multiplier, "cost_multiplier")
    if cost_multiplier < 0:
        raise ValueError("cost_multiplier must not be negative")

    paths: Dict[str, SignalBacktestResult] = {}
    for factor_name, spec in scenario.signals.items():
        if spec.weight == 0.0:
            continue
        paths[factor_name] = _validated_path(provider.backtest(spec), factor_name)
    if not paths:
        raise ValueError(f"Scenario '{scenario.name}' has no non-zero signals")

    all_dates = sorted(set().union(*(set(path.holdings.index) for path in paths.values())))
    all_assets = sorted(set().union(*(set(path.holdings.columns) for path in paths.values())))
    index = pd.DatetimeIndex(all_dates)
    weighted_holdings: Dict[str, pd.DataFrame] = {}
    weighted_gross_pnl: Dict[str, pd.DataFrame] = {}
    for factor_name, path in paths.items():
        scale = scenario.scaler * scenario.signals[factor_name].weight
        weighted_holdings[factor_name] = (
            path.holdings.reindex(index=index, columns=all_assets).fillna(0.0) * scale
        )
        weighted_gross_pnl[factor_name] = (
            path.gross_asset_pnl.reindex(index=index, columns=all_assets).fillna(0.0)
            * scale
        )

    gross_asset_pnl = sum(weighted_gross_pnl.values())
    aggregate_holdings = sum(weighted_holdings.values())
    signal_costs: Dict[str, pd.DataFrame] = {
        factor_name: pd.DataFrame(0.0, index=index, columns=all_assets)
        for factor_name in paths
    }

    if cost_mode == "sleeve":
        trade_volume = pd.DataFrame(0.0, index=index, columns=all_assets)
        gross_exposure = pd.DataFrame(0.0, index=index, columns=all_assets)
        for factor_name, holdings in weighted_holdings.items():
            path = paths[factor_name]
            trades = holdings.diff().abs().fillna(0.0)
            rates = path.cost_rates.reindex(all_assets).fillna(DEFAULT_COST_RATE)
            signal_costs[factor_name] = trades.multiply(rates * cost_multiplier)
            trade_volume = trade_volume.add(trades, fill_value=0.0)
            gross_exposure = gross_exposure.add(holdings.abs(), fill_value=0.0)
    else:
        trade_volume = pd.DataFrame(0.0, index=index, columns=all_assets)
        gross_exposure = pd.DataFrame(0.0, index=index, columns=all_assets)
        buckets = sorted({path.execution_bucket for path in paths.values()})
        for bucket in buckets:
            names = [name for name, path in paths.items() if path.execution_bucket == bucket]
            bucket_holdings = sum(weighted_holdings[name] for name in names)
            bucket_trades = bucket_holdings.diff().abs().fillna(0.0)
            trade_volume = trade_volume.add(bucket_trades, fill_value=0.0)
            gross_exposure = gross_exposure.add(bucket_holdings.abs(), fill_value=0.0)

            rates = pd.Series(index=all_assets, dtype=float)
            for asset in all_assets:
                observed = {
                    float(paths[name].cost_rates.get(asset, DEFAULT_COST_RATE))
                    for name in names
                    if asset in paths[name].holdings.columns
                }
                if len(observed) > 1:
                    raise ValueError(
                        f"Execution bucket '{bucket}' has inconsistent cost rates "
                        f"for asset '{asset}'"
                    )
                rates.loc[asset] = next(iter(observed), DEFAULT_COST_RATE)
            bucket_cost = bucket_trades.multiply(rates * cost_multiplier)

            individual_trades = {
                name: weighted_holdings[name].diff().abs().fillna(0.0) for name in names
            }
            allocation_base = sum(individual_trades.values()).replace(0.0, np.nan)
            for name in names:
                allocation = individual_trades[name].div(allocation_base)
                signal_costs[name] = signal_costs[name].add(
                    bucket_cost.multiply(allocation).fillna(0.0), fill_value=0.0
                )

    costs_by_asset = sum(signal_costs.values())
    net_asset_pnl = gross_asset_pnl.subtract(costs_by_asset, fill_value=0.0)
    signal_pnl = pd.DataFrame(
        {
            factor_name: weighted_gross_pnl[factor_name].subtract(
                signal_costs[factor_name], fill_value=0.0
            ).sum(axis=1)
            for factor_name in paths
        },
        index=index,
    )
    signal_asset_pnl = {
        factor_name: weighted_gross_pnl[factor_name].subtract(
            signal_costs[factor_name], fill_value=0.0
        )
        for factor_name in paths
    }
    return PortfolioResult(
        scenario=scenario,
        cost_mode=cost_mode,
        signal_source=str(getattr(provider, "signal_source", "provided")),
        gross_asset_pnl=gross_asset_pnl,
        costs_by_asset=costs_by_asset,
        net_asset_pnl=net_asset_pnl,
        aggregate_holdings=aggregate_holdings,
        trade_volume=trade_volume,
        gross_exposure=gross_exposure,
        signal_pnl=signal_pnl,
        signal_asset_pnl=signal_asset_pnl,
    )


def _safe_ratio(numerator: float, denominator: float) -> float:
    if denominator == 0 or not math.isfinite(denominator):
        return np.nan
    return numerator / denominator


def _metric_series(
    pnl: pd.Series,
    trade_volume: pd.Series,
    exposure: pd.Series,
    *,
    business_days_per_year: int,
) -> pd.Series:
    pnl = pnl.fillna(0.0)
    daily_std = float(pnl.std())
    daily_mean = float(pnl.mean())
    cumulative = pnl.cumsum()
    drawdown = cumulative - cumulative.cummax().clip(lower=0.0)
    average_trade = float(trade_volume.fillna(0.0).mean())
    average_exposure = float(exposure.fillna(0.0).mean())
    return pd.Series(
        {
            "annualized_pnl": daily_mean * business_days_per_year,
            "daily_std": daily_std,
            "sharpe": _safe_ratio(
                daily_mean * math.sqrt(business_days_per_year), daily_std
            ),
            "max_drawdown": float(drawdown.min()) if len(drawdown) else np.nan,
            "turnover_pct": 100.0 * _safe_ratio(average_trade, average_exposure),
            "pnl_per_turnover_bps": 10000.0
            * _safe_ratio(daily_mean, average_trade),
            "total_pnl": float(pnl.sum()),
            "positive_day_pct": 100.0 * float((pnl > 0).mean()),
        }
    )


def portfolio_metrics(
    result: PortfolioResult,
    *,
    business_days_per_year: int = BUSINESS_DAYS_PER_YEAR,
) -> pd.Series:
    return _metric_series(
        result.portfolio_pnl,
        result.trade_volume.sum(axis=1),
        result.gross_exposure.sum(axis=1),
        business_days_per_year=business_days_per_year,
    )


def asset_metrics(
    result: PortfolioResult,
    *,
    business_days_per_year: int = BUSINESS_DAYS_PER_YEAR,
) -> pd.DataFrame:
    rows = {}
    for asset in result.net_asset_pnl.columns:
        rows[asset] = _metric_series(
            result.net_asset_pnl[asset],
            result.trade_volume[asset],
            result.gross_exposure[asset],
            business_days_per_year=business_days_per_year,
        )
    return pd.DataFrame(rows).T


def _tenor_start(end_date: pd.Timestamp, tenor: str) -> pd.Timestamp:
    match = re.fullmatch(r"([1-9][0-9]*)([my])", str(tenor).strip().lower())
    if not match:
        raise ValueError(f"Invalid PNL tenor '{tenor}'; use values such as 3m or 2y")
    count = int(match.group(1))
    if match.group(2) == "m":
        return end_date - pd.DateOffset(months=count)
    return end_date - pd.DateOffset(years=count)


def _precomputed_btmetrics_fallback(
    result: PortfolioResult,
    tenors: Sequence[str],
    business_days_per_year: int,
) -> Mapping[str, Any]:
    """Dependency-light equivalent of MetricsBase.calculate_pnl_stats_from_pnl."""

    asset_pnl = result.net_asset_pnl.fillna(0.0)
    portfolio_pnl = asset_pnl.sum(axis=1)
    performance: dict[str, pd.Series] = {}
    for metric in ("sharpe", "std"):
        values: dict[str, float] = {}
        samples = {"": portfolio_pnl}
        end_date = portfolio_pnl.index[-1]
        samples.update(
            {
                f"_{tenor}": portfolio_pnl.loc[_tenor_start(end_date, tenor) :]
                for tenor in tenors
            }
        )
        for suffix, sample in samples.items():
            value = float(sample.std())
            if metric == "sharpe":
                value = _safe_ratio(
                    float(sample.mean()) * math.sqrt(business_days_per_year), value
                )
            values[f"{metric}{suffix}"] = value
        performance[metric] = pd.Series(values)

    trades = result.aggregate_holdings.diff().abs()
    avg_trades = trades.mean()
    avg_exposure = result.aggregate_holdings.abs().mean()
    return {
        **performance,
        "turnover": 100.0 * avg_trades / avg_exposure,
        "pnl_per_trade": 10000.0 * asset_pnl.mean() / avg_trades,
    }


def btmetrics_result(
    result: PortfolioResult,
    *,
    tenors: Sequence[str] = DEFAULT_PNL_TENORS,
    business_days_per_year: int = BUSINESS_DAYS_PER_YEAR,
) -> BtMetricsResult:
    """Return the notebook's btmetrics views for one aggregated portfolio.

    Portfolio PNL already includes execution slippage and portfolio-level cost
    netting, so it is passed to ``MetricsBase`` as precomputed asset PNL rather
    than being reconstructed from synthetic returns.
    """

    normalized_tenors = tuple(str(tenor).strip().lower() for tenor in tenors)
    if result.net_asset_pnl.empty:
        raise ValueError(f"Scenario '{result.scenario.name}' has no portfolio PNL")
    for tenor in normalized_tenors:
        _tenor_start(result.net_asset_pnl.index[-1], tenor)

    try:
        from pycmqlib3.analytics.btmetrics import MetricsBase

        zero_returns = pd.DataFrame(
            0.0,
            index=result.aggregate_holdings.index,
            columns=result.aggregate_holdings.columns,
        )
        calculator = MetricsBase(
            holdings=result.aggregate_holdings.copy(),
            returns=zero_returns,
            shift_holdings=0,
            cost_dict={},
            business_days_per_year=business_days_per_year,
        )
        stats = calculator.calculate_pnl_stats_from_pnl(
            result.net_asset_pnl,
            holdings=result.aggregate_holdings,
            tenors=list(normalized_tenors),
            perf_metrics=["sharpe", "std"],
        )
        engine = "pycmqlib3.analytics.btmetrics.MetricsBase"
    except ImportError:
        stats = _precomputed_btmetrics_fallback(
            result, normalized_tenors, business_days_per_year
        )
        engine = "btmetrics-compatible fallback"

    portfolio_rows = []
    for tenor in ("full", *normalized_tenors):
        suffix = "" if tenor == "full" else f"_{tenor}"
        portfolio_rows.append(
            {
                "tenor": tenor,
                "sharpe": float(stats["sharpe"].get(f"sharpe{suffix}", np.nan)),
                "std": float(stats["std"].get(f"std{suffix}", np.nan)),
            }
        )
    portfolio = pd.DataFrame(portfolio_rows).set_index("tenor")
    assets = pd.concat(
        [
            pd.Series(stats["turnover"], name="turnover"),
            pd.Series(stats["pnl_per_trade"], name="pnl_per_trade"),
        ],
        axis=1,
    ).reindex(result.net_asset_pnl.columns)
    assets.index.name = "asset"
    return BtMetricsResult(
        portfolio=portfolio,
        assets=assets,
        tenors=normalized_tenors,
        engine=engine,
    )


def _comparison_frame(current: pd.DataFrame, proposed: pd.DataFrame) -> pd.DataFrame:
    index = current.index.union(proposed.index)
    columns = current.columns.union(proposed.columns)
    before = current.reindex(index=index, columns=columns)
    after = proposed.reindex(index=index, columns=columns)
    rows = []
    for asset in index:
        for metric in columns:
            current_value = before.loc[asset, metric]
            proposed_value = after.loc[asset, metric]
            rows.append(
                {
                    "asset": asset,
                    "metric": metric,
                    "current": current_value,
                    "proposed": proposed_value,
                    "delta": proposed_value - current_value,
                }
            )
    return pd.DataFrame(rows).set_index(["asset", "metric"])


def compare_portfolios(
    baseline: PortfolioResult,
    proposed: PortfolioResult,
    *,
    business_days_per_year: int = BUSINESS_DAYS_PER_YEAR,
    pnl_tenors: Sequence[str] = DEFAULT_PNL_TENORS,
) -> ScenarioComparison:
    """Create before/after/delta tables and validation checks."""

    current_metrics = portfolio_metrics(
        baseline, business_days_per_year=business_days_per_year
    )
    proposed_metrics = portfolio_metrics(
        proposed, business_days_per_year=business_days_per_year
    )
    summary = pd.DataFrame(
        {
            "current": current_metrics,
            "proposed": proposed_metrics,
            "delta": proposed_metrics - current_metrics,
        }
    )

    current_assets = asset_metrics(
        baseline, business_days_per_year=business_days_per_year
    )
    proposed_assets = asset_metrics(
        proposed, business_days_per_year=business_days_per_year
    )
    asset_comparison = _comparison_frame(current_assets, proposed_assets)
    current_btmetrics = btmetrics_result(
        baseline,
        tenors=pnl_tenors,
        business_days_per_year=business_days_per_year,
    )
    proposed_btmetrics = btmetrics_result(
        proposed,
        tenors=pnl_tenors,
        business_days_per_year=business_days_per_year,
    )

    factor_names = sorted(
        set(baseline.scenario.signals).union(proposed.scenario.signals)
    )
    current_signal_pnl = baseline.signal_pnl.reindex(columns=factor_names).fillna(0.0)
    proposed_signal_pnl = proposed.signal_pnl.reindex(columns=factor_names).fillna(0.0)
    signal_rows = []
    for factor_name in factor_names:
        old_spec = baseline.scenario.signals.get(factor_name)
        new_spec = proposed.scenario.signals.get(factor_name)
        old_weight = old_spec.weight if old_spec else 0.0
        new_weight = new_spec.weight if new_spec else 0.0
        if old_weight == 0 and new_weight != 0:
            status = "added"
        elif old_weight != 0 and new_weight == 0:
            status = "removed"
        elif old_spec and new_spec and old_spec.cache_key() != new_spec.cache_key():
            status = "definition changed"
        elif old_weight != new_weight:
            status = "weight changed"
        else:
            status = "unchanged"
        old_pnl = float(current_signal_pnl[factor_name].sum())
        new_pnl = float(proposed_signal_pnl[factor_name].sum())
        signal_rows.append(
            {
                "factor_name": factor_name,
                "signal_name_current": old_spec.name if old_spec else "",
                "signal_name_proposed": new_spec.name if new_spec else "",
                "status": status,
                "weight_current": old_weight,
                "weight_proposed": new_weight,
                "weight_delta": new_weight - old_weight,
                "total_pnl_current": old_pnl,
                "total_pnl_proposed": new_pnl,
                "total_pnl_delta": new_pnl - old_pnl,
            }
        )
    signal_attribution = pd.DataFrame(signal_rows).set_index("factor_name")

    daily_pnl = pd.concat(
        [baseline.portfolio_pnl, proposed.portfolio_pnl], axis=1
    ).fillna(0.0)
    daily_pnl.columns = ["current", "proposed"]
    daily_pnl["delta"] = daily_pnl["proposed"] - daily_pnl["current"]
    daily_pnl["current_cumulative"] = daily_pnl["current"].cumsum()
    daily_pnl["proposed_cumulative"] = daily_pnl["proposed"].cumsum()
    daily_pnl["delta_cumulative"] = daily_pnl["delta"].cumsum()

    delta_std = float(daily_pnl["delta"].std())
    info_ratio = _safe_ratio(
        float(daily_pnl["delta"].mean()) * math.sqrt(business_days_per_year),
        delta_std,
    )
    summary.loc["tracking_error_annualized", :] = [
        np.nan,
        np.nan,
        delta_std * math.sqrt(business_days_per_year),
    ]
    summary.loc["information_ratio", :] = [np.nan, np.nan, info_ratio]
    summary.loc["pnl_correlation", :] = [
        np.nan,
        np.nan,
        daily_pnl[["current", "proposed"]].corr().iloc[0, 1],
    ]

    pnl_reconciliation_current = float(
        (baseline.signal_pnl.sum(axis=1) - baseline.portfolio_pnl).abs().max()
    )
    pnl_reconciliation_proposed = float(
        (proposed.signal_pnl.sum(axis=1) - proposed.portfolio_pnl).abs().max()
    )
    checks = pd.DataFrame(
        [
            {
                "check": "current signal PNL reconciles to portfolio",
                "actual": pnl_reconciliation_current,
                "tolerance": 1e-8,
                "status": "OK" if pnl_reconciliation_current <= 1e-8 else "FAIL",
            },
            {
                "check": "proposed signal PNL reconciles to portfolio",
                "actual": pnl_reconciliation_proposed,
                "tolerance": 1e-8,
                "status": "OK" if pnl_reconciliation_proposed <= 1e-8 else "FAIL",
            },
            {
                "check": "comparison has overlapping dates",
                "actual": len(daily_pnl),
                "tolerance": 1,
                "status": "OK" if len(daily_pnl) >= 1 else "FAIL",
            },
        ]
    ).set_index("check")

    return ScenarioComparison(
        baseline=baseline,
        proposed=proposed,
        summary=summary,
        asset_comparison=asset_comparison,
        signal_attribution=signal_attribution,
        daily_pnl=daily_pnl,
        checks=checks,
        current_btmetrics=current_btmetrics,
        proposed_btmetrics=proposed_btmetrics,
    )


def run_strategy_comparison(
    baseline: StrategyScenario,
    proposed: StrategyScenario,
    provider: SignalProvider,
    *,
    cost_mode: str = "netted",
    cost_multiplier: float = 1.0,
    pnl_tenors: Sequence[str] = DEFAULT_PNL_TENORS,
) -> ScenarioComparison:
    current_result = compose_portfolio(
        baseline,
        provider,
        cost_mode=cost_mode,
        cost_multiplier=cost_multiplier,
    )
    proposed_result = compose_portfolio(
        proposed,
        provider,
        cost_mode=cost_mode,
        cost_multiplier=cost_multiplier,
    )
    return compare_portfolios(
        current_result, proposed_result, pnl_tenors=pnl_tenors
    )


def _strategy_assets(scenario: StrategyScenario) -> list[str]:
    from pycmqlib3.utility.misc import inst2product

    assets = scenario.config.get("assets", [])
    products = []
    for item in assets:
        if not isinstance(item, dict) or not item.get("underliers"):
            continue
        product = inst2product(str(item["underliers"][0]))
        if product not in products:
            products.append(product)
    if not products:
        raise ValueError(f"No assets found in {scenario.strategy_file}")
    return products


def _continuous_price_frame(
    futures: pd.DataFrame,
    assets: Sequence[str],
    window: str,
) -> pd.DataFrame:
    """Extract a continuous-price window using the notebook fallback chain."""

    fallback_map = {
        "close": ["close"],
        "d_twap": ["d_twap", "close"],
        "a1535": ["a1535", "d_twap", "close"],
        "a1505": ["a1505", "a1535", "d_twap", "close"],
        "n305": ["n305", "a1505", "a1535", "d_twap", "close"],
        "n310": ["n310", "a1505", "a1535", "d_twap", "close"],
        "n315": ["n315", "a1505", "a1535", "d_twap", "close"],
        "n450": ["n450", "a1505", "a1535", "d_twap", "close"],
    }
    candidates = fallback_map.get(window, [window])
    output: Dict[str, pd.Series] = {}
    missing: list[str] = []
    for asset in assets:
        continuous = f"{asset}c1"
        series: pd.Series | None = None
        for candidate in candidates:
            column = (continuous, candidate)
            if column not in futures.columns:
                continue
            candidate_series = pd.to_numeric(futures[column], errors="coerce")
            series = candidate_series if series is None else series.combine_first(candidate_series)
        if series is None:
            missing.append(continuous)
        else:
            output[asset] = series
    if missing:
        raise ValueError(
            f"Missing continuous '{window}' prices (including fallbacks) for: {missing}"
        )
    frame = pd.DataFrame(output, index=futures.index)
    frame.index = pd.to_datetime(frame.index)
    return frame.sort_index().ffill()


def _load_saved_daily_futures(as_of: dt.date) -> pd.DataFrame:
    """Read the daily parquet cache without importing the WtPy update stack."""

    data_dir = Path("C:/dev/data")
    requested = data_dir / f"fut_d_{as_of:%Y%m%d}.parquet"
    if requested.exists():
        return pd.read_parquet(requested)
    candidates = sorted(data_dir.glob("fut_d_*.parquet"), reverse=True)
    cutoff = as_of.strftime("%Y%m%d")
    eligible = [path for path in candidates if path.stem.rsplit("_", 1)[-1] <= cutoff]
    if not eligible:
        raise ValueError(f"Saved futures data is unavailable as of {as_of}")
    return pd.read_parquet(eligible[0])


def build_production_factor_provider(
    scenarios: Sequence[StrategyScenario],
    *,
    start_date: dt.date,
    end_date: dt.date,
    as_of: dt.date | None = None,
    holding_lag: int = 2,
    default_cost_rate: float = DEFAULT_COST_RATE,
) -> FactorFrameSignalProvider:
    """Load factor DB data and saved continuous prices for scenario backtests.

    All scenarios must describe the same strategy universe and data settings.
    This function performs external data reads but does not update prices, the
    database, the source workbook, or strategy JSON files.
    """

    if not scenarios:
        raise ValueError("At least one scenario is required")
    reference = scenarios[0]
    assets = _strategy_assets(reference)
    for scenario in scenarios[1:]:
        if _strategy_assets(scenario) != assets:
            raise ValueError("Scenarios use different asset universes")

    config = reference.config
    roll_label = str(config.get("roll_label", "hot"))
    freq = str(config.get("freq", "d1"))
    db_table = str(config.get("fact_db_table", "fut_fact_data"))
    vol_key = str(config.get("vol_key", "pct_vol"))
    repo_type = str(config.get("repo_type", "asset"))
    if any(str(item.config.get("repo_type", "asset")) != repo_type for item in scenarios):
        raise ValueError("Scenarios use different repo_type values")

    factor_names = sorted(
        {
            spec.name
            for scenario in scenarios
            for spec in scenario.signals.values()
            if spec.weight != 0.0
        }
    )
    warmup_start = start_date - dt.timedelta(days=730)
    from pycmqlib3.strategy.signal_repo import signal_execution_config
    from pycmqlib3.utility.dbaccess import load_factor_data

    query_products = [] if repo_type == "port" else assets
    factor_long = load_factor_data(
        query_products,
        factor_list=factor_names,
        roll_label=roll_label,
        start=warmup_start,
        end=end_date,
        freq=freq,
        db_table=db_table,
    )
    vol_long = load_factor_data(
        assets,
        factor_list=[vol_key],
        roll_label=roll_label,
        start=warmup_start,
        end=end_date,
        freq=freq,
        db_table=db_table,
    )
    if factor_long.empty:
        raise ValueError("Factor database query returned no signal data")
    if vol_long.empty:
        raise ValueError(f"Factor database query returned no '{vol_key}' data")

    factor_long = factor_long.copy()
    factor_long["date"] = pd.to_datetime(factor_long["date"])
    factor_values: Dict[str, pd.DataFrame] = {}
    for factor_name in factor_names:
        subset = factor_long[factor_long["fact_name"] == factor_name]
        if subset.empty:
            continue
        if repo_type == "port":
            series = subset.groupby("date")["fact_val"].last()
            factor_values[factor_name] = pd.concat(
                [series.rename(asset) for asset in assets], axis=1
            )
        else:
            frame = subset.pivot_table(
                index="date",
                columns="product_code",
                values="fact_val",
                aggfunc="last",
            )
            factor_values[factor_name] = frame.reindex(columns=assets)

    vol_long = vol_long.copy()
    vol_long["date"] = pd.to_datetime(vol_long["date"])
    volatility = vol_long.pivot_table(
        index="date",
        columns="product_code",
        values="fact_val",
        aggfunc="last",
    ).reindex(columns=assets)

    as_of = end_date if as_of is None else as_of
    futures = _load_saved_daily_futures(as_of)
    if futures is None or futures.empty:
        raise ValueError(f"Saved futures data is unavailable as of {as_of}")
    close_prices = _continuous_price_frame(futures, assets, "close")
    execution_names = {
        _execution_signal_name(spec)
        for scenario in scenarios
        for spec in scenario.signals.values()
        if spec.weight != 0.0
    }
    required_windows = {
        str(signal_execution_config.get(name, {"win": "n305"})["win"])
        for name in execution_names
    }
    execution_prices = {
        window: _continuous_price_frame(futures, assets, window)
        for window in sorted(required_windows)
    }
    asset_returns = close_prices.pct_change().loc[
        pd.Timestamp(warmup_start) : pd.Timestamp(end_date)
    ]
    volatility = volatility.reindex(index=asset_returns.index).ffill()
    factor_values = {
        name: frame.reindex(index=asset_returns.index).ffill()
        for name, frame in factor_values.items()
    }
    cost_rates = {
        asset: BOND_COST_RATE if asset in {"T", "TF", "TS", "TL"} else default_cost_rate
        for asset in assets
    }
    provider = FactorFrameSignalProvider(
        factor_values=factor_values,
        asset_returns=asset_returns,
        volatility=volatility,
        cost_rates=cost_rates,
        holding_lag=holding_lag,
        close_prices=close_prices.reindex(index=asset_returns.index).ffill(),
        execution_prices={
            window: frame.reindex(index=asset_returns.index).ffill()
            for window, frame in execution_prices.items()
        },
        execution_config=signal_execution_config,
        default_execution_window="n305",
    )
    provider.analysis_start = pd.Timestamp(start_date)  # type: ignore[attr-defined]
    provider.signal_source = "factor_db"  # type: ignore[attr-defined]
    return provider


def build_generated_historical_provider(
    scenarios: Sequence[StrategyScenario],
    *,
    start_date: dt.date,
    end_date: dt.date,
    as_of: dt.date | None = None,
    holding_lag: int = 2,
    default_cost_rate: float = DEFAULT_COST_RATE,
    price_loader: Callable[[dt.date], pd.DataFrame] | None = None,
    fundamental_loader: Callable[[dt.date], pd.DataFrame] | None = None,
    metrics_class: type | None = None,
) -> FactorFrameSignalProvider:
    """Generate configured signals from historical spot and futures data.

    This mirrors the signal-construction section of
    ``bktest/bktest_prod_daily_run.ipynb``: historical futures and ``spot_df``
    are loaded as of the requested date, price-derived features are added,
    signal names are routed through the production registries, and recipes are
    evaluated from ``signal_store``.  It never reads precomputed factor values
    from ``fut_fact_data``.
    """

    if not scenarios:
        raise ValueError("At least one scenario is required")
    reference = scenarios[0]
    assets = _strategy_assets(reference)
    for scenario in scenarios[1:]:
        if _strategy_assets(scenario) != assets:
            raise ValueError("Scenarios use different asset universes")

    from misc_scripts.fun_factor_update import (
        factors_by_asset,
        factors_by_beta_neutral,
        factors_by_func,
        factors_by_spread,
        factors_by_spread2,
        single_factors,
        spread_config,
    )
    from misc_scripts.update_fut_prices import load_fun_data
    from pycmqlib3.strategy.signal_repo import (
        commod_phycarry_dict,
        get_funda_signal_from_store,
        signal_execution_config,
        signal_store,
    )
    from pycmqlib3.utility.exch_ctd_func import (
        SH_ctd_basis,
        io_ctd_basis,
        lc_ctd_basis,
        si_ctd_basis,
    )
    from misc_scripts.historical_signal_generator import (
        HistoricalSignalGenerator,
        prepare_historical_feature_data,
    )

    as_of = end_date if as_of is None else as_of
    if as_of < end_date:
        raise ValueError("as_of must be on or after end_date")
    if price_loader is None:
        price_loader = _load_saved_daily_futures
    if fundamental_loader is None:
        fundamental_loader = load_fun_data

    futures = price_loader(as_of)
    if futures is None or futures.empty:
        raise ValueError(f"Saved futures data is unavailable as of {as_of}")
    futures = futures.copy().sort_index()
    futures.index = pd.to_datetime(futures.index)
    futures = futures.loc[: pd.Timestamp(end_date)]

    spot_df = fundamental_loader(as_of)
    if spot_df is None or spot_df.empty:
        raise ValueError(f"Historical spot_df is unavailable as of {as_of}")
    spot_df = spot_df.copy().sort_index()
    spot_df.index = pd.to_datetime(spot_df.index)
    spot_df = spot_df.loc[: pd.Timestamp(end_date)]

    ctd_features = {
        "io_ctd_spot": ("i", io_ctd_basis),
        "si_ctd_spot": ("si", si_ctd_basis),
        "SH_ctd_spot": ("SH", SH_ctd_basis),
        "lc_ctd_spot": ("lc", lc_ctd_basis),
    }
    for feature_name, (asset, function) in ctd_features.items():
        expiry_column = (f"{asset}c1", "expiry")
        if expiry_column in futures.columns:
            spot_df[feature_name] = function(spot_df, futures[expiry_column])

    spot_df = prepare_historical_feature_data(
        futures,
        spot_df,
        assets,
        commod_phycarry_dict=commod_phycarry_dict,
    )
    specs = [
        spec
        for scenario in scenarios
        for spec in scenario.signals.values()
        if spec.weight != 0.0
    ]
    if not specs:
        raise ValueError("Scenarios contain no non-zero signals")

    vol_window = int(reference.config.get("vol_win", 20))
    generator = HistoricalSignalGenerator(
        futures,
        spot_df,
        assets,
        signal_store=signal_store,
        get_signal=get_funda_signal_from_store,
        factors_by_asset=factors_by_asset,
        single_factors=single_factors,
        factors_by_spread=factors_by_spread,
        factors_by_spread2=factors_by_spread2,
        factors_by_beta_neutral=factors_by_beta_neutral,
        factors_by_func=factors_by_func,
        spread_config=spread_config,
        execution_name=_execution_signal_name,
        vol_window=vol_window,
    )
    bundle = generator.generate(specs)

    close_prices = _continuous_price_frame(futures, assets, "close")
    asset_returns = close_prices.pct_change()
    volatility = asset_returns.rolling(vol_window).std()
    execution_config = {
        name: dict(settings) for name, settings in signal_execution_config.items()
    }
    execution_config.update(
        {name: dict(settings) for name, settings in bundle.execution_overrides.items()}
    )
    execution_names = {_execution_signal_name(spec) for spec in specs}
    required_windows = {
        str(execution_config.get(name, {"win": "n305"})["win"])
        for name in execution_names
    }
    execution_prices = {
        window: _continuous_price_frame(futures, assets, window)
        for window in sorted(required_windows)
    }

    warmup_start = pd.Timestamp(start_date) - pd.Timedelta(days=730)
    history_index = asset_returns.loc[warmup_start : pd.Timestamp(end_date)].index
    asset_returns = asset_returns.reindex(history_index)
    volatility = volatility.reindex(history_index).ffill()
    factor_values = {
        name: frame.reindex(history_index).ffill()
        for name, frame in bundle.factor_values.items()
    }
    cost_rates = {
        asset: BOND_COST_RATE if asset in {"T", "TF", "TS", "TL"} else default_cost_rate
        for asset in assets
    }
    provider = FactorFrameSignalProvider(
        factor_values=factor_values,
        asset_returns=asset_returns,
        volatility=volatility,
        cost_rates=cost_rates,
        holding_lag=holding_lag,
        close_prices=close_prices.reindex(history_index).ffill(),
        execution_prices={
            window: frame.reindex(history_index).ffill()
            for window, frame in execution_prices.items()
        },
        execution_config=execution_config,
        volatility_overrides={
            name: frame.reindex(history_index).ffill()
            for name, frame in bundle.volatility_overrides.items()
        },
        traded_price_overrides={
            name: frame.reindex(history_index).ffill()
            for name, frame in bundle.traded_price_overrides.items()
        },
        pnl_modes=bundle.pnl_modes,
        post_funcs=bundle.post_funcs,
        default_execution_window="n305",
        metrics_class=metrics_class,
    )
    provider.analysis_start = pd.Timestamp(start_date)  # type: ignore[attr-defined]
    provider.signal_routes = dict(bundle.routes)  # type: ignore[attr-defined]
    provider.signal_source = "historical_generation"  # type: ignore[attr-defined]
    return provider


def trim_comparison_start(
    comparison: ScenarioComparison, start_date: str | dt.date | pd.Timestamp
) -> ScenarioComparison:
    """Recompute a comparison after excluding the provider warm-up period."""

    cutoff = pd.Timestamp(start_date)

    def trim(result: PortfolioResult) -> PortfolioResult:
        return replace(
            result,
            gross_asset_pnl=result.gross_asset_pnl.loc[cutoff:],
            costs_by_asset=result.costs_by_asset.loc[cutoff:],
            net_asset_pnl=result.net_asset_pnl.loc[cutoff:],
            aggregate_holdings=result.aggregate_holdings.loc[cutoff:],
            trade_volume=result.trade_volume.loc[cutoff:],
            gross_exposure=result.gross_exposure.loc[cutoff:],
            signal_pnl=result.signal_pnl.loc[cutoff:],
            signal_asset_pnl={
                name: frame.loc[cutoff:]
                for name, frame in result.signal_asset_pnl.items()
            },
        )

    return compare_portfolios(
        trim(comparison.baseline),
        trim(comparison.proposed),
        pnl_tenors=comparison.current_btmetrics.tenors,
    )


def _write_dataframe_sheet(
    workbook: Workbook,
    sheet_name: str,
    frame: pd.DataFrame,
    *,
    table_name: str,
) -> Any:
    worksheet = workbook.create_sheet(sheet_name)
    worksheet.sheet_view.showGridLines = False
    worksheet.freeze_panes = "A2"
    output = frame.copy()
    if isinstance(output.columns, pd.MultiIndex):
        output.columns = ["_".join(str(item) for item in col) for col in output.columns]
    index_names = list(output.index.names)
    if any(name is None for name in index_names):
        index_names = [name or "index" for name in index_names]
        output.index.names = index_names
    output = output.reset_index()
    worksheet.append(list(output.columns))
    for row in output.itertuples(index=False, name=None):
        worksheet.append(
            [
                value.to_pydatetime() if isinstance(value, pd.Timestamp) else value
                for value in row
            ]
        )

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    for column_index, column_name in enumerate(output.columns, start=1):
        values = [str(column_name)] + [str(value) for value in output.iloc[:, column_index - 1].head(100)]
        worksheet.column_dimensions[get_column_letter(column_index)].width = min(
            max(max(len(value) for value in values) + 2, 12), 32
        )
        if pd.api.types.is_datetime64_any_dtype(output.iloc[:, column_index - 1]):
            for cell in worksheet.iter_cols(
                min_col=column_index,
                max_col=column_index,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                for item in cell:
                    item.number_format = "yyyy-mm-dd"
        elif pd.api.types.is_numeric_dtype(output.iloc[:, column_index - 1]):
            for cell in worksheet.iter_cols(
                min_col=column_index,
                max_col=column_index,
                min_row=2,
                max_row=worksheet.max_row,
            ):
                for item in cell:
                    item.number_format = "#,##0.0000;[Red](#,##0.0000);-"

    if worksheet.max_row >= 2:
        table = Table(
            displayName=table_name,
            ref=f"A1:{get_column_letter(worksheet.max_column)}{worksheet.max_row}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    return worksheet


def write_comparison_excel(
    comparison: ScenarioComparison,
    excel_path: str | Path,
) -> Path:
    """Write a formatted, auditable comparison workbook."""

    workbook = Workbook()
    workbook.remove(workbook.active)
    summary_sheet = _write_dataframe_sheet(
        workbook, "Summary", comparison.summary, table_name="PortfolioSummary"
    )
    _write_dataframe_sheet(
        workbook,
        "Current Portfolio",
        comparison.current_btmetrics.portfolio,
        table_name="CurrentPortfolioBtMetrics",
    )
    _write_dataframe_sheet(
        workbook,
        "Proposed Portfolio",
        comparison.proposed_btmetrics.portfolio,
        table_name="ProposedPortfolioBtMetrics",
    )
    _write_dataframe_sheet(
        workbook,
        "Current Assets",
        comparison.current_btmetrics.assets,
        table_name="CurrentAssetBtMetrics",
    )
    _write_dataframe_sheet(
        workbook,
        "Proposed Assets",
        comparison.proposed_btmetrics.assets,
        table_name="ProposedAssetBtMetrics",
    )
    asset_sheet = _write_dataframe_sheet(
        workbook,
        "Asset Comparison",
        comparison.asset_comparison,
        table_name="AssetComparison",
    )
    signal_sheet = _write_dataframe_sheet(
        workbook,
        "Signal Attribution",
        comparison.signal_attribution,
        table_name="SignalAttribution",
    )
    _write_dataframe_sheet(
        workbook,
        "Current Signal PNL",
        comparison.baseline.signal_pnl,
        table_name="CurrentSignalPnl",
    )
    _write_dataframe_sheet(
        workbook,
        "Proposed Signal PNL",
        comparison.proposed.signal_pnl,
        table_name="ProposedSignalPnl",
    )
    _write_dataframe_sheet(
        workbook,
        "Current Signal Metrics",
        signal_asset_diagnostics(comparison.baseline),
        table_name="CurrentSignalMetrics",
    )
    _write_dataframe_sheet(
        workbook,
        "Proposed Signal Metrics",
        signal_asset_diagnostics(comparison.proposed),
        table_name="ProposedSignalMetrics",
    )
    _write_dataframe_sheet(
        workbook,
        "Signal Coverage",
        signal_coverage(comparison),
        table_name="SignalCoverage",
    )
    daily_output = comparison.daily_pnl.copy()
    daily_output["chart_date"] = daily_output.index.strftime("%Y-%m-%d")
    daily_sheet = _write_dataframe_sheet(
        workbook, "Daily PNL", daily_output, table_name="DailyPnl"
    )
    checks_sheet = _write_dataframe_sheet(
        workbook, "Checks", comparison.checks, table_name="BacktestChecks"
    )
    run_info = pd.DataFrame(
        {
            "value": [
                comparison.baseline.scenario.strategy_file,
                str(comparison.baseline.scenario.source_path or ""),
                comparison.baseline.scenario.scaler,
                comparison.proposed.scenario.scaler,
                comparison.baseline.cost_mode,
                comparison.baseline.signal_source,
                comparison.daily_pnl.index.min().strftime("%Y-%m-%d"),
                comparison.daily_pnl.index.max().strftime("%Y-%m-%d"),
                len(comparison.daily_pnl),
                BUSINESS_DAYS_PER_YEAR,
                ", ".join(comparison.current_btmetrics.tenors),
                comparison.current_btmetrics.engine,
                comparison.proposed_btmetrics.engine,
                "Normalized risk-unit PNL. Standard deviation is daily; Sharpe is annualized with 244 business days; turnover is percent and pnl_per_trade is basis points.",
                "exec_assets is treated as an exclusion list, matching current production position code.",
            ]
        },
        index=[
            "strategy_file",
            "source_json",
            "current_scaler",
            "proposed_scaler",
            "cost_mode",
            "signal_source",
            "start_date",
            "end_date",
            "observations",
            "business_days_per_year",
            "pnl_tenors",
            "current_metrics_engine",
            "proposed_metrics_engine",
            "metric_units",
            "exec_assets_semantics",
        ],
    )
    _write_dataframe_sheet(workbook, "Run Info", run_info, table_name="RunInformation")

    def add_delta_rules(worksheet: Any, header_names: Iterable[str]) -> None:
        header_map = {
            worksheet.cell(1, column).value: column
            for column in range(1, worksheet.max_column + 1)
        }
        for header_name in header_names:
            if header_name not in header_map or worksheet.max_row < 2:
                continue
            letter = get_column_letter(header_map[header_name])
            target = f"{letter}2:{letter}{worksheet.max_row}"
            worksheet.conditional_formatting.add(
                target,
                CellIsRule(
                    operator="greaterThan",
                    formula=["0"],
                    fill=PatternFill("solid", fgColor="C6EFCE"),
                ),
            )
            worksheet.conditional_formatting.add(
                target,
                CellIsRule(
                    operator="lessThan",
                    formula=["0"],
                    fill=PatternFill("solid", fgColor="FFC7CE"),
                ),
            )

    add_delta_rules(summary_sheet, ["delta"])
    add_delta_rules(asset_sheet, ["delta"])
    add_delta_rules(signal_sheet, ["weight_delta", "total_pnl_delta"])

    for row in range(2, checks_sheet.max_row + 1):
        status_cell = checks_sheet.cell(row, checks_sheet.max_column)
        status_cell.fill = PatternFill(
            "solid", fgColor="C6EFCE" if status_cell.value == "OK" else "FFC7CE"
        )

    chart = LineChart()
    chart.title = "Current vs Proposed Cumulative PNL"
    chart.y_axis.title = "Cumulative PNL"
    chart.x_axis.title = "Date"
    chart.height = 9
    chart.width = 18
    header_map = {
        daily_sheet.cell(1, column).value: column
        for column in range(1, daily_sheet.max_column + 1)
    }
    chart.add_data(
        Reference(
            daily_sheet,
            min_col=header_map["current_cumulative"],
            max_col=header_map["proposed_cumulative"],
            min_row=1,
            max_row=daily_sheet.max_row,
        ),
        titles_from_data=True,
    )
    chart.series[0].tx = SeriesLabel(v="Current")
    chart.series[1].tx = SeriesLabel(v="Proposed")
    chart.set_categories(
        Reference(
            daily_sheet,
            min_col=header_map["chart_date"],
            min_row=2,
            max_row=daily_sheet.max_row,
        )
    )
    chart.x_axis.tickLblSkip = max(1, daily_sheet.max_row // 12)
    summary_sheet.add_chart(chart, "F2")

    checks_sheet["F1"] = "Model status"
    checks_sheet["F1"].font = Font(bold=True)
    checks_sheet["F2"] = (
        "OK" if all(comparison.checks["status"] == "OK") else "REVIEW"
    )
    checks_sheet["F2"].fill = PatternFill(
        "solid",
        fgColor="C6EFCE" if checks_sheet["F2"].value == "OK" else "FFC7CE",
    )

    excel_path = Path(excel_path)
    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(excel_path)
    workbook.close()
    return excel_path.resolve()


def _format_report_value(metric: str, value: Any, *, signed: bool = False) -> str:
    if value is None or pd.isna(value):
        return "—"
    number = float(value)
    sign = "+" if signed and number > 0 else ""
    if metric in {"turnover_pct", "positive_day_pct"}:
        return f"{sign}{number:,.2f}%"
    if metric in {"sharpe", "pnl_correlation", "information_ratio"}:
        return f"{sign}{number:,.2f}"
    if metric == "pnl_per_turnover_bps":
        return f"{sign}{number:,.2f} bps"
    return f"{sign}{number:,.4f}"


def _report_metric_label(metric: str) -> str:
    labels = {
        "annualized_pnl": "Annualized P&L",
        "daily_std": "Daily standard deviation",
        "annualized_std": "Annualized standard deviation",
        "sharpe": "Sharpe ratio",
        "max_drawdown": "Maximum drawdown",
        "turnover_pct": "Turnover",
        "pnl_per_turnover_bps": "P&L per turnover",
        "total_pnl": "Total P&L",
        "positive_day_pct": "Positive days",
        "tracking_error_annualized": "Annualized tracking error",
        "information_ratio": "Information ratio",
        "pnl_correlation": "P&L correlation",
    }
    return labels.get(metric, metric.replace("_", " ").title())


def _performance_table_html(summary: pd.DataFrame) -> str:
    rows = []
    for metric, values in summary.iterrows():
        rows.append(
            {
                "Metric": _report_metric_label(str(metric)),
                "Current": _format_report_value(str(metric), values.get("current")),
                "Proposed": _format_report_value(str(metric), values.get("proposed")),
                "Change": _format_report_value(
                    str(metric), values.get("delta"), signed=True
                ),
            }
        )
    return pd.DataFrame(rows).to_html(
        index=False,
        classes="report-table performance-table",
        border=0,
        escape=True,
    )


def _btmetrics_table_html(result: BtMetricsResult, table: str) -> str:
    if table == "portfolio":
        output = result.portfolio.reset_index()
        output.columns = ["Tenor", "Sharpe", "Daily std"]
        for column in ("Sharpe", "Daily std"):
            output[column] = output[column].map(
                lambda value: "—" if pd.isna(value) else f"{float(value):,.4f}"
            )
    elif table == "assets":
        output = result.assets.reset_index()
        output.columns = ["Asset", "Turnover", "PNL per trade"]
        output["Turnover"] = output["Turnover"].map(
            lambda value: "—" if pd.isna(value) else f"{float(value):,.2f}%"
        )
        output["PNL per trade"] = output["PNL per trade"].map(
            lambda value: "—" if pd.isna(value) else f"{float(value):,.2f} bps"
        )
    else:
        raise ValueError("table must be 'portfolio' or 'assets'")
    return output.to_html(
        index=False,
        classes="report-table btmetrics-table",
        border=0,
        escape=True,
    )


def _signal_coverage_table_html(comparison: ScenarioComparison) -> str:
    coverage = signal_coverage(comparison).reset_index()
    columns = [
        "signal",
        "first_active_current",
        "last_active_current",
        "active_days_current",
        "total_pnl_current",
        "first_active_proposed",
        "last_active_proposed",
        "active_days_proposed",
        "total_pnl_proposed",
    ]
    output = coverage.loc[:, columns].copy()
    output.columns = [
        "Signal",
        "Current first active",
        "Current last active",
        "Current active days",
        "Current total PNL",
        "Proposed first active",
        "Proposed last active",
        "Proposed active days",
        "Proposed total PNL",
    ]
    for column in (
        "Current first active",
        "Current last active",
        "Proposed first active",
        "Proposed last active",
    ):
        output[column] = output[column].map(
            lambda value: "—" if pd.isna(value) else pd.Timestamp(value).strftime("%Y-%m-%d")
        )
    for column in ("Current total PNL", "Proposed total PNL"):
        output[column] = output[column].map(lambda value: f"{float(value):,.4f}")
    return output.to_html(
        index=False,
        classes="report-table signal-coverage-table",
        border=0,
        escape=True,
    )


def _signal_change_table_html(attribution: pd.DataFrame) -> str:
    changed = attribution[attribution["status"] != "unchanged"].copy()
    if changed.empty:
        return '<p class="empty-state">No signal definitions or weights changed.</p>'
    changed = changed.reset_index()
    columns = [
        "factor_name",
        "status",
        "weight_current",
        "weight_proposed",
        "weight_delta",
        "total_pnl_delta",
    ]
    changed = changed.loc[:, columns]
    changed.columns = [
        "Factor",
        "Status",
        "Current weight",
        "Proposed weight",
        "Weight change",
        "P&L change",
    ]
    for column in ["Current weight", "Proposed weight", "Weight change", "P&L change"]:
        changed[column] = changed[column].map(
            lambda value: "—" if pd.isna(value) else f"{float(value):+,.4f}"
        )
    return changed.to_html(
        index=False,
        classes="report-table signal-table",
        border=0,
        escape=True,
    )


def _comparison_chart_html(comparison: ScenarioComparison) -> tuple[str, str, str]:
    try:
        import plotly.graph_objects as go
    except ImportError as exc:
        raise ImportError(
            "HTML chart generation requires plotly, which is already used by "
            "pycmqlib3.analytics.tstool in the production environment."
        ) from exc

    daily = comparison.daily_pnl.copy()
    dates = daily.index
    current_cumulative = daily["current"].cumsum()
    proposed_cumulative = daily["proposed"].cumsum()
    current_drawdown = current_cumulative - current_cumulative.cummax()
    proposed_drawdown = proposed_cumulative - proposed_cumulative.cummax()
    current_name = "Current"
    proposed_name = "Proposed"
    blue = "#2F6BFF"
    gold = "#C58A13"

    base_layout = {
        "template": "plotly_white",
        "height": 410,
        "margin": dict(l=58, r=24, t=58, b=48),
        "font": dict(family="Inter, Segoe UI, Arial, sans-serif", color="#27344D"),
        "paper_bgcolor": "rgba(0,0,0,0)",
        "plot_bgcolor": "rgba(0,0,0,0)",
        "hovermode": "x unified",
        "legend": dict(orientation="h", y=1.13, x=0),
        "xaxis": dict(showgrid=False),
        "yaxis": dict(gridcolor="#E7EBF2", zerolinecolor="#AAB4C5"),
    }

    cumulative = go.Figure()
    cumulative.add_trace(
        go.Scatter(
            x=dates,
            y=current_cumulative,
            name=current_name,
            mode="lines",
            line=dict(color=blue, width=2.4),
        )
    )
    cumulative.add_trace(
        go.Scatter(
            x=dates,
            y=proposed_cumulative,
            name=proposed_name,
            mode="lines",
            line=dict(color=gold, width=2.4, dash="dash"),
        )
    )
    cumulative.update_layout(title="Cumulative portfolio P&L", **base_layout)

    drawdown = go.Figure()
    drawdown.add_trace(
        go.Scatter(
            x=dates,
            y=current_drawdown,
            name=current_name,
            mode="lines",
            line=dict(color=blue, width=2.1),
        )
    )
    drawdown.add_trace(
        go.Scatter(
            x=dates,
            y=proposed_drawdown,
            name=proposed_name,
            mode="lines",
            line=dict(color=gold, width=2.1, dash="dash"),
        )
    )
    drawdown.update_layout(title="Portfolio drawdown", **base_layout)

    distribution = go.Figure()
    distribution.add_trace(
        go.Histogram(
            x=daily["current"],
            name=current_name,
            marker=dict(color=blue, line=dict(color="#1D459F", width=0.5)),
            opacity=0.62,
            histnorm="probability",
        )
    )
    distribution.add_trace(
        go.Histogram(
            x=daily["proposed"],
            name=proposed_name,
            marker=dict(color=gold, line=dict(color="#75510B", width=0.5)),
            opacity=0.55,
            histnorm="probability",
        )
    )
    distribution_layout = dict(base_layout)
    distribution_layout["hovermode"] = "closest"
    distribution.update_layout(
        title="Daily portfolio P&L distribution",
        barmode="overlay",
        xaxis_title="Daily normalized P&L",
        yaxis_title="Share of observations",
        **distribution_layout,
    )

    config = {"responsive": True, "displaylogo": False}
    return (
        cumulative.to_html(
            full_html=False, include_plotlyjs=True, config=config
        ),
        drawdown.to_html(full_html=False, include_plotlyjs=False, config=config),
        distribution.to_html(
            full_html=False, include_plotlyjs=False, config=config
        ),
    )


def _btmetrics_chart_html(comparison: ScenarioComparison) -> tuple[str, str]:
    """Render notebook-style interactive charts for the btmetrics tables."""

    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError as exc:
        raise ImportError(
            "HTML btmetrics charts require plotly, matching the notebook's iplot output."
        ) from exc

    blue = "#2F6BFF"
    gold = "#C58A13"
    current = comparison.current_btmetrics.portfolio
    proposed = comparison.proposed_btmetrics.portfolio
    tenors = list(current.index)
    portfolio = make_subplots(
        rows=1,
        cols=2,
        subplot_titles=("Sharpe ratio by tenor", "Daily standard deviation by tenor"),
        horizontal_spacing=0.12,
    )
    for column, col_num in (("sharpe", 1), ("std", 2)):
        portfolio.add_trace(
            go.Scatter(
                x=tenors,
                y=current[column],
                name="Current",
                legendgroup="current",
                showlegend=col_num == 1,
                mode="lines+markers",
                line=dict(color=blue, width=2.4),
            ),
            row=1,
            col=col_num,
        )
        portfolio.add_trace(
            go.Scatter(
                x=tenors,
                y=proposed.reindex(tenors)[column],
                name="Proposed",
                legendgroup="proposed",
                showlegend=col_num == 1,
                mode="lines+markers",
                line=dict(color=gold, width=2.4, dash="dash"),
            ),
            row=1,
            col=col_num,
        )
    portfolio.update_layout(
        template="plotly_white",
        height=430,
        margin=dict(l=58, r=24, t=70, b=48),
        font=dict(family="Inter, Segoe UI, Arial, sans-serif", color="#27344D"),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        hovermode="x unified",
        legend=dict(orientation="h", y=1.16, x=0),
    )
    portfolio.update_xaxes(type="category", showgrid=False)
    portfolio.update_yaxes(gridcolor="#E7EBF2", zerolinecolor="#AAB4C5")

    current_assets = comparison.current_btmetrics.assets
    proposed_assets = comparison.proposed_btmetrics.assets
    assets = list(current_assets.index.union(proposed_assets.index))
    efficiency = make_subplots(
        rows=1,
        cols=2,
        subplot_titles=("Asset turnover", "PNL per trade"),
        horizontal_spacing=0.12,
    )
    for column, col_num in (("turnover", 1), ("pnl_per_trade", 2)):
        efficiency.add_trace(
            go.Bar(
                x=assets,
                y=current_assets.reindex(assets)[column],
                name="Current",
                legendgroup="current",
                showlegend=col_num == 1,
                marker_color=blue,
                opacity=0.82,
            ),
            row=1,
            col=col_num,
        )
        efficiency.add_trace(
            go.Bar(
                x=assets,
                y=proposed_assets.reindex(assets)[column],
                name="Proposed",
                legendgroup="proposed",
                showlegend=col_num == 1,
                marker_color=gold,
                opacity=0.72,
            ),
            row=1,
            col=col_num,
        )
    efficiency.update_layout(
        template="plotly_white",
        height=470,
        barmode="group",
        margin=dict(l=58, r=24, t=70, b=80),
        font=dict(family="Inter, Segoe UI, Arial, sans-serif", color="#27344D"),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        legend=dict(orientation="h", y=1.16, x=0),
    )
    efficiency.update_xaxes(type="category", tickangle=-35, showgrid=False)
    efficiency.update_yaxes(gridcolor="#E7EBF2", zerolinecolor="#AAB4C5")
    efficiency.update_yaxes(title_text="Percent", row=1, col=1)
    efficiency.update_yaxes(title_text="Basis points", row=1, col=2)

    config = {"responsive": True, "displaylogo": False}
    return (
        portfolio.to_html(full_html=False, include_plotlyjs=False, config=config),
        efficiency.to_html(full_html=False, include_plotlyjs=False, config=config),
    )


def signal_coverage(comparison: ScenarioComparison) -> pd.DataFrame:
    """Summarize active history and PNL for every signal in both scenarios."""

    names = sorted(
        set(comparison.baseline.signal_pnl.columns).union(
            comparison.proposed.signal_pnl.columns
        )
    )
    rows = []
    for name in names:
        row: dict[str, Any] = {"signal": name}
        for label, result in (
            ("current", comparison.baseline),
            ("proposed", comparison.proposed),
        ):
            pnl = result.signal_pnl.reindex(columns=[name], fill_value=0.0)[name]
            active = pnl.abs() > 1e-12
            row[f"first_active_{label}"] = pnl.index[active].min() if active.any() else pd.NaT
            row[f"last_active_{label}"] = pnl.index[active].max() if active.any() else pd.NaT
            row[f"active_days_{label}"] = int(active.sum())
            row[f"total_pnl_{label}"] = float(pnl.sum())
        rows.append(row)
    return pd.DataFrame(rows).set_index("signal")


def signal_asset_diagnostics(
    result: PortfolioResult,
    *,
    business_days_per_year: int = BUSINESS_DAYS_PER_YEAR,
) -> pd.DataFrame:
    """Return signal- and asset-level diagnostics from net daily PNL."""

    rows: list[dict[str, Any]] = []
    for signal_name, asset_pnl in result.signal_asset_pnl.items():
        asset_pnl = asset_pnl.fillna(0.0).sort_index()
        series_by_label = {"Total": asset_pnl.sum(axis=1)}
        series_by_label.update(
            {str(asset): asset_pnl[asset] for asset in asset_pnl.columns}
        )
        for asset, pnl in series_by_label.items():
            pnl = pnl.fillna(0.0)
            daily_mean = float(pnl.mean())
            daily_std = float(pnl.std())
            cumulative = pnl.cumsum()
            drawdown = cumulative - cumulative.cummax().clip(lower=0.0)
            active = pnl.ne(0.0)
            rows.append(
                {
                    "signal": signal_name,
                    "asset": asset,
                    "sharpe": _safe_ratio(
                        daily_mean * math.sqrt(business_days_per_year), daily_std
                    ),
                    "daily_std": daily_std,
                    "annualized_pnl": daily_mean * business_days_per_year,
                    "total_pnl": float(pnl.sum()),
                    "max_drawdown": float(drawdown.min()) if len(drawdown) else np.nan,
                    "positive_day_pct": 100.0 * float((pnl > 0.0).mean()),
                    "active_days": int(active.sum()),
                    "first_active": active[active].index[0] if active.any() else pd.NaT,
                    "last_active": active[active].index[-1] if active.any() else pd.NaT,
                }
            )
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows).set_index(["signal", "asset"]).sort_index()


def _signal_chart_html(comparison: ScenarioComparison) -> tuple[str, str]:
    """Render interactive per-signal paths and total contribution charts."""

    try:
        import plotly.graph_objects as go
    except ImportError as exc:
        raise ImportError("HTML signal charts require plotly/iplot support") from exc

    current = comparison.baseline.signal_pnl.fillna(0.0)
    proposed = comparison.proposed.signal_pnl.fillna(0.0)
    names = list(current.columns.union(proposed.columns))
    totals = pd.DataFrame(
        {
            "current": current.reindex(columns=names, fill_value=0.0).sum(),
            "proposed": proposed.reindex(columns=names, fill_value=0.0).sum(),
        }
    )
    order = (
        totals.abs().max(axis=1).sort_values(ascending=False).index.tolist()
    )
    blue = "#2F6BFF"
    gold = "#C58A13"
    paths = go.Figure()
    for position, name in enumerate(order):
        current_path = current.reindex(columns=[name], fill_value=0.0)[name].cumsum()
        proposed_path = proposed.reindex(columns=[name], fill_value=0.0)[name].cumsum()
        paths.add_trace(
            go.Scatter(
                x=current_path.index,
                y=current_path,
                name="Current",
                legendgroup="current",
                mode="lines",
                visible=position == 0,
                line=dict(color=blue, width=2.4),
            )
        )
        paths.add_trace(
            go.Scatter(
                x=proposed_path.index,
                y=proposed_path,
                name="Proposed",
                legendgroup="proposed",
                mode="lines",
                visible=position == 0,
                line=dict(color=gold, width=2.4, dash="dash"),
            )
        )
    buttons = []
    for position, name in enumerate(order):
        visible = [False] * (2 * len(order))
        visible[2 * position] = True
        visible[2 * position + 1] = True
        buttons.append(
            {
                "label": name,
                "method": "update",
                "args": [
                    {"visible": visible},
                    {"title": f"Cumulative PNL by signal — {name}"},
                ],
            }
        )
    paths.update_layout(
        title=f"Cumulative PNL by signal — {order[0] if order else ''}",
        template="plotly_white",
        height=470,
        margin=dict(l=58, r=24, t=115, b=48),
        font=dict(family="Inter, Segoe UI, Arial, sans-serif", color="#27344D"),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        hovermode="x unified",
        legend=dict(orientation="h", y=1.08, x=0),
        updatemenus=[
            {
                "buttons": buttons,
                "direction": "down",
                "showactive": True,
                "x": 1,
                "y": 1.20,
                "xanchor": "right",
                "yanchor": "top",
            }
        ],
    )
    paths.update_xaxes(showgrid=False)
    paths.update_yaxes(gridcolor="#E7EBF2", zerolinecolor="#AAB4C5")

    contributions = go.Figure()
    contributions.add_trace(
        go.Bar(
            x=order,
            y=totals.reindex(order)["current"],
            name="Current",
            marker_color=blue,
            opacity=0.82,
        )
    )
    contributions.add_trace(
        go.Bar(
            x=order,
            y=totals.reindex(order)["proposed"],
            name="Proposed",
            marker_color=gold,
            opacity=0.72,
        )
    )
    contributions.update_layout(
        title="Total PNL contribution by signal",
        template="plotly_white",
        height=520,
        barmode="group",
        margin=dict(l=58, r=24, t=70, b=170),
        font=dict(family="Inter, Segoe UI, Arial, sans-serif", color="#27344D"),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        legend=dict(orientation="h", y=1.12, x=0),
    )
    contributions.update_xaxes(type="category", tickangle=-55, showgrid=False)
    contributions.update_yaxes(gridcolor="#E7EBF2", zerolinecolor="#AAB4C5")
    config = {"responsive": True, "displaylogo": False}
    return (
        paths.to_html(full_html=False, include_plotlyjs=False, config=config),
        contributions.to_html(
            full_html=False, include_plotlyjs=False, config=config
        ),
    )


def _signal_asset_chart_html(result: PortfolioResult, label: str) -> str:
    """Render selectable cumulative asset PNL and diagnostics for each signal."""

    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError as exc:
        raise ImportError("HTML signal charts require plotly/iplot support") from exc

    diagnostics = signal_asset_diagnostics(result)
    signals = list(result.signal_asset_pnl)
    if not signals:
        return '<p class="empty-state">No non-zero signals are available.</p>'

    figure = make_subplots(
        rows=2,
        cols=1,
        specs=[[{"type": "xy"}], [{"type": "table"}]],
        row_heights=[0.61, 0.39],
        vertical_spacing=0.08,
        subplot_titles=("Cumulative net PNL by asset", "Signal and asset metrics"),
    )
    palette = [
        "#2F6BFF",
        "#C58A13",
        "#D46A2E",
        "#6F7F31",
        "#C4547D",
        "#667085",
        "#7057C7",
        "#138A8A",
        "#8C5A3C",
    ]
    trace_groups: dict[str, list[int]] = {}
    for signal_position, signal_name in enumerate(signals):
        visible = signal_position == 0
        asset_pnl = result.signal_asset_pnl[signal_name].fillna(0.0).sort_index()
        active_assets = [
            str(asset) for asset in asset_pnl.columns if asset_pnl[asset].ne(0.0).any()
        ]
        trace_groups[signal_name] = []
        series = [("Total", asset_pnl.sum(axis=1))] + [
            (asset, asset_pnl[asset]) for asset in active_assets
        ]
        for series_position, (asset, pnl) in enumerate(series):
            trace_groups[signal_name].append(len(figure.data))
            figure.add_trace(
                go.Scatter(
                    x=pnl.index,
                    y=pnl.cumsum(),
                    name=asset,
                    mode="lines",
                    visible=visible,
                    line=dict(
                        color="#17233B"
                        if asset == "Total"
                        else palette[(series_position - 1) % len(palette)],
                        width=3.0 if asset == "Total" else 1.5,
                        dash="solid" if asset == "Total" else "dot",
                    ),
                    hovertemplate=f"{asset}<br>%{{x|%Y-%m-%d}}<br>%{{y:,.2f}}<extra></extra>",
                ),
                row=1,
                col=1,
            )

        signal_metrics = diagnostics.xs(signal_name).reindex(
            ["Total"] + active_assets
        )
        trace_groups[signal_name].append(len(figure.data))
        figure.add_trace(
            go.Table(
                visible=visible,
                columnwidth=[75, 70, 90, 105, 105, 105, 90, 80],
                header=dict(
                    values=[
                        "Asset",
                        "Sharpe",
                        "Daily std",
                        "Annualized PNL",
                        "Total PNL",
                        "Max drawdown",
                        "Positive days",
                        "Active days",
                    ],
                    fill_color="#EDF2FB",
                    align="left",
                    font=dict(color="#36445E", size=12),
                    height=30,
                ),
                cells=dict(
                    values=[
                        list(signal_metrics.index),
                        ["—" if pd.isna(v) else f"{v:,.3f}" for v in signal_metrics["sharpe"]],
                        [f"{v:,.2f}" for v in signal_metrics["daily_std"]],
                        [f"{v:,.2f}" for v in signal_metrics["annualized_pnl"]],
                        [f"{v:,.2f}" for v in signal_metrics["total_pnl"]],
                        [f"{v:,.2f}" for v in signal_metrics["max_drawdown"]],
                        [f"{v:,.2f}%" for v in signal_metrics["positive_day_pct"]],
                        [f"{int(v):,}" for v in signal_metrics["active_days"]],
                    ],
                    fill_color=[
                        ["#F7F9FD"] + ["#FFFFFF"] * max(0, len(signal_metrics) - 1)
                    ],
                    align="left",
                    font=dict(color="#27344D", size=11),
                    height=27,
                ),
            ),
            row=2,
            col=1,
        )

    buttons = []
    for signal_name in signals:
        visible = [False] * len(figure.data)
        for trace_index in trace_groups[signal_name]:
            visible[trace_index] = True
        buttons.append(
            {
                "label": signal_name,
                "method": "update",
                "args": [
                    {"visible": visible},
                    {"title": f"{label} signal backtest — {signal_name}"},
                ],
            }
        )

    figure.update_layout(
        title=f"{label} signal backtest — {signals[0]}",
        template="plotly_white",
        height=850,
        margin=dict(l=58, r=24, t=120, b=30),
        font=dict(family="Inter, Segoe UI, Arial, sans-serif", color="#27344D"),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        hovermode="x unified",
        legend=dict(orientation="h", y=1.07, x=0),
        updatemenus=[
            {
                "buttons": buttons,
                "direction": "down",
                "showactive": True,
                "x": 1,
                "y": 1.16,
                "xanchor": "right",
                "yanchor": "top",
            }
        ],
    )
    figure.update_xaxes(showgrid=False, row=1, col=1)
    figure.update_yaxes(
        title_text="Cumulative PNL",
        gridcolor="#E7EBF2",
        zerolinecolor="#AAB4C5",
        row=1,
        col=1,
    )
    return figure.to_html(
        full_html=False,
        include_plotlyjs=False,
        config={"responsive": True, "displaylogo": False},
    )


def write_comparison_html(
    comparison: ScenarioComparison,
    html_path: str | Path,
) -> Path:
    """Write a self-contained portfolio before/after HTML report."""

    cumulative_chart, drawdown_chart, distribution_chart = _comparison_chart_html(
        comparison
    )
    signal_path_chart, signal_contribution_chart = _signal_chart_html(comparison)
    current_signal_asset_chart = _signal_asset_chart_html(
        comparison.baseline, "Current"
    )
    proposed_signal_asset_chart = _signal_asset_chart_html(
        comparison.proposed, "Proposed"
    )
    summary = comparison.summary
    headline_metrics = ["sharpe", "daily_std", "turnover_pct", "total_pnl"]
    cards = []
    for metric in headline_metrics:
        values = summary.loc[metric]
        cards.append(
            f"""
            <article class="metric-card">
              <div class="metric-label">{html_lib.escape(_report_metric_label(metric))}</div>
              <div class="metric-value">{html_lib.escape(_format_report_value(metric, values['proposed']))}</div>
              <div class="metric-context">
                Current {html_lib.escape(_format_report_value(metric, values['current']))}
                <span class="delta">Change {html_lib.escape(_format_report_value(metric, values['delta'], signed=True))}</span>
              </div>
            </article>
            """
        )

    sharpe = summary.loc["sharpe"]
    risk = summary.loc["daily_std"]
    turnover = summary.loc["turnover_pct"]
    correlation = summary.loc["pnl_correlation", "delta"]
    start = comparison.daily_pnl.index.min().strftime("%d %b %Y")
    end = comparison.daily_pnl.index.max().strftime("%d %b %Y")
    title = "Portfolio Backtest Comparison"
    strategy = html_lib.escape(comparison.baseline.scenario.strategy_file)
    performance_table = _performance_table_html(summary)
    current_portfolio_table = _btmetrics_table_html(
        comparison.current_btmetrics, "portfolio"
    )
    proposed_portfolio_table = _btmetrics_table_html(
        comparison.proposed_btmetrics, "portfolio"
    )
    current_asset_table = _btmetrics_table_html(
        comparison.current_btmetrics, "assets"
    )
    proposed_asset_table = _btmetrics_table_html(
        comparison.proposed_btmetrics, "assets"
    )
    signal_coverage_table = _signal_coverage_table_html(comparison)
    signal_table = _signal_change_table_html(comparison.signal_attribution)
    document = f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <meta name="color-scheme" content="light dark">
  <title>{title}</title>
  <style>
    :root {{ color-scheme: light; --paper:#f5f7fb; --card:#fff; --ink:#17233b; --muted:#657087; --line:#dfe5ee; --blue:#2f6bff; --gold:#c58a13; }}
    * {{ box-sizing:border-box; }}
    body {{ margin:0; background:var(--paper); color:var(--ink); font-family:Inter,"Segoe UI",Arial,sans-serif; line-height:1.5; }}
    main {{ width:min(1180px, calc(100% - 32px)); margin:0 auto; padding:42px 0 64px; }}
    header {{ margin-bottom:26px; }}
    h1 {{ margin:0 0 8px; font-size:clamp(28px,4vw,44px); letter-spacing:-.035em; }}
    h2 {{ margin:0 0 10px; font-size:24px; letter-spacing:-.02em; }}
    p {{ margin:7px 0; }}
    .eyebrow {{ color:var(--blue); font-size:13px; font-weight:700; letter-spacing:.09em; text-transform:uppercase; }}
    .scope {{ color:var(--muted); }}
    .section {{ background:var(--card); border:1px solid var(--line); border-radius:16px; padding:24px; margin:18px 0; box-shadow:0 8px 30px rgba(30,52,89,.06); }}
    .summary-list {{ margin:10px 0 0; padding-left:22px; }}
    .summary-list li {{ margin:9px 0; }}
    .metric-grid {{ display:grid; grid-template-columns:repeat(4,minmax(0,1fr)); gap:14px; margin:18px 0; }}
    .metric-card {{ background:var(--card); border:1px solid var(--line); border-radius:14px; padding:18px; }}
    .metric-label {{ color:var(--muted); font-size:13px; font-weight:650; }}
    .metric-value {{ font-size:28px; font-weight:750; margin:5px 0; font-variant-numeric:tabular-nums; }}
    .metric-context {{ color:var(--muted); font-size:12px; display:flex; flex-wrap:wrap; gap:8px; }}
    .delta {{ color:var(--gold); font-weight:700; }}
    .chart-wrap {{ width:100%; min-height:410px; }}
    .signal-chart-wrap {{ width:100%; min-height:850px; }}
    .caption {{ color:var(--muted); font-size:14px; margin-bottom:10px; }}
    .table-scroll {{ overflow-x:auto; }}
    .split-grid {{ display:grid; grid-template-columns:repeat(2,minmax(0,1fr)); gap:18px; }}
    .split-grid h3 {{ margin:0 0 8px; }}
    .report-table {{ width:100%; border-collapse:collapse; font-size:14px; font-variant-numeric:tabular-nums; }}
    .report-table th {{ text-align:left; background:#edf2fb; color:#36445e; padding:11px 12px; border-bottom:1px solid var(--line); white-space:nowrap; }}
    .report-table td {{ padding:10px 12px; border-bottom:1px solid var(--line); white-space:nowrap; }}
    .report-table tbody tr:hover {{ background:#f7f9fd; }}
    .empty-state {{ color:var(--muted); font-style:italic; }}
    .note {{ border-left:4px solid var(--gold); padding:4px 0 4px 14px; color:var(--muted); }}
    .report-tabs {{ display:flex; gap:8px; margin:0 0 18px; padding:6px; width:max-content; max-width:100%; background:#e9eef7; border-radius:12px; }}
    .tab-button {{ border:0; border-radius:9px; padding:10px 16px; background:transparent; color:var(--muted); font:inherit; font-weight:700; cursor:pointer; }}
    .tab-button.active {{ background:var(--card); color:var(--ink); box-shadow:0 2px 8px rgba(30,52,89,.10); }}
    .tab-panel {{ display:none; }}
    .tab-panel.active {{ display:block; }}
    footer {{ color:var(--muted); font-size:12px; margin-top:22px; }}
    @media (max-width:850px) {{ .metric-grid {{ grid-template-columns:repeat(2,minmax(0,1fr)); }} .split-grid {{ grid-template-columns:1fr; }} main {{ width:min(100% - 20px,1180px); padding-top:24px; }} .section {{ padding:18px; }} }}
    @media (max-width:480px) {{ .metric-grid {{ grid-template-columns:1fr; }} }}
    @media print {{ body {{ background:#fff; }} main {{ width:100%; padding:0; }} .section,.metric-card {{ box-shadow:none; break-inside:avoid; }} .report-tabs {{ display:none; }} .tab-panel {{ display:block; }} }}
  </style>
</head>
<body>
<main>
  <header>
    <div class="eyebrow">Strategy impact analysis</div>
    <h1>{title}</h1>
    <p class="scope">{strategy} · {start} to {end} · {len(comparison.daily_pnl):,} observations</p>
  </header>

  <nav class="report-tabs" aria-label="Report sections">
    <button class="tab-button active" type="button" data-tab="portfolio-panel">Portfolio summary</button>
    <button class="tab-button" type="button" data-tab="signals-panel">Signal diagnostics</button>
  </nav>

  <div id="portfolio-panel" class="tab-panel active">

  <section class="section">
    <h2>Executive Summary</h2>
    <ul class="summary-list">
      <li><strong>Risk-adjusted performance:</strong> Sharpe moved from {_format_report_value('sharpe', sharpe['current'])} to {_format_report_value('sharpe', sharpe['proposed'])}, a change of {_format_report_value('sharpe', sharpe['delta'], signed=True)}.</li>
      <li><strong>Portfolio risk:</strong> daily standard deviation moved from {_format_report_value('daily_std', risk['current'])} to {_format_report_value('daily_std', risk['proposed'])}. Sharpe remains annualized using {BUSINESS_DAYS_PER_YEAR} business days.</li>
      <li><strong>Trading intensity:</strong> turnover moved from {_format_report_value('turnover_pct', turnover['current'])} to {_format_report_value('turnover_pct', turnover['proposed'])}; daily P&L correlation is {_format_report_value('pnl_correlation', correlation)}.</li>
    </ul>
  </section>

  <section aria-label="Headline portfolio metrics" class="metric-grid">{''.join(cards)}</section>

  <section class="section">
    <h2>Portfolio growth paths</h2>
    <p class="caption">Cumulative normalized P&L shows when the proposed construction diverges from the current portfolio. Solid blue is current; dashed gold is proposed.</p>
    <div class="chart-wrap">{cumulative_chart}</div>
  </section>

  <section class="section">
    <h2>Downside path and recovery</h2>
    <p class="caption">Drawdown compares peak-to-trough loss and recovery timing. A deeper negative value indicates a larger historical loss from the prior P&L peak.</p>
    <div class="chart-wrap">{drawdown_chart}</div>
  </section>

  <section class="section">
    <h2>Daily outcome distribution</h2>
    <p class="caption">The overlapping distributions show whether the proposal changes the spread or tails of daily portfolio outcomes, beyond the headline volatility number.</p>
    <div class="chart-wrap">{distribution_chart}</div>
  </section>

  <section class="section">
    <h2>Portfolio performance table</h2>
    <p class="caption">Exact current, proposed, and signed-change values for the full comparison period.</p>
    <div class="table-scroll">{performance_table}</div>
  </section>

  <section class="section">
    <h2>Portfolio btmetrics by tenor</h2>
    <p class="caption">The same Sharpe and daily standard-deviation definitions used by the notebook, calculated separately for the current and proposed portfolios.</p>
    <div class="split-grid">
      <div><h3>Current portfolio</h3><div class="table-scroll">{current_portfolio_table}</div></div>
      <div><h3>Proposed portfolio</h3><div class="table-scroll">{proposed_portfolio_table}</div></div>
    </div>
  </section>

  <section class="section">
    <h2>Asset trading efficiency</h2>
    <p class="caption">Asset turnover and PNL per trade use the btmetrics formulas on each fully aggregated portfolio.</p>
    <div class="split-grid">
      <div><h3>Current portfolio</h3><div class="table-scroll">{current_asset_table}</div></div>
      <div><h3>Proposed portfolio</h3><div class="table-scroll">{proposed_asset_table}</div></div>
    </div>
  </section>

  </div>

  <div id="signals-panel" class="tab-panel">
  <section class="section">
    <h2>PNL by signal</h2>
    <p class="caption">Select a signal from the dropdown to inspect its current and proposed cumulative PNL paths. The contribution chart ranks signals by absolute portfolio impact.</p>
    <div class="chart-wrap">{signal_path_chart}</div>
    <div class="chart-wrap">{signal_contribution_chart}</div>
  </section>

  <section class="section">
    <h2>Current signal backtests by asset</h2>
    <p class="caption">Choose a signal to review cumulative net PNL for its total sleeve and each traded asset. The synchronized table reports signal Sharpe and asset Sharpe alongside daily standard deviation, annualized PNL, drawdown, positive days, and active days.</p>
    <div class="signal-chart-wrap">{current_signal_asset_chart}</div>
  </section>

  <section class="section">
    <h2>Proposed signal backtests by asset</h2>
    <p class="caption">The same diagnostics after applying the proposed signal configuration and weights.</p>
    <div class="signal-chart-wrap">{proposed_signal_asset_chart}</div>
  </section>

  <section class="section">
    <h2>Signal data coverage</h2>
    <p class="caption">First and last non-zero PNL dates make truncated or unavailable factor histories explicit.</p>
    <div class="table-scroll">{signal_coverage_table}</div>
  </section>

  </div>

  <section class="section">
    <h2>Signals changed in the proposal</h2>
    <p class="caption">Weight and total P&L attribution for added, removed, or modified signals.</p>
    <div class="table-scroll">{signal_table}</div>
  </section>

  <section class="section">
    <h2>Decision notes and assumptions</h2>
    <p class="note">Execution prices and lags follow <code>signal_execution_config</code> for the corresponding production signal name. For example, <code>a1505</code> and <code>n305</code> are distinct point-in-time execution buckets and are not interchangeable.</p>
    <p>Signal source: <code>{html_lib.escape(comparison.baseline.signal_source)}</code>. Signal P&L uses <code>btmetrics.MetricsBase</code>; continuous-return signals use <code>calculate_daily_pnl</code>, while spread-contract price-difference signals use <code>calculate_pnl_stats</code>, matching the production notebook. After aggregation, precomputed net asset P&L is evaluated through <code>calculate_pnl_stats_from_pnl</code>, so execution slippage and portfolio-level cost netting are retained. Transaction costs are applied by the scenario portfolio layer so offsetting trades are netted only when they share an execution bucket.</p>
    <p>Results are normalized risk-unit P&L rather than currency P&L. Historical performance is descriptive and does not remove the need to validate signal data availability at each execution point.</p>
  </section>

  <footer>Generated from the strategy scenario comparison output.</footer>
</main>
<script>
  document.querySelectorAll('.tab-button').forEach((button) => {{
    button.addEventListener('click', () => {{
      document.querySelectorAll('.tab-button').forEach((item) => item.classList.remove('active'));
      document.querySelectorAll('.tab-panel').forEach((panel) => panel.classList.remove('active'));
      button.classList.add('active');
      document.getElementById(button.dataset.tab).classList.add('active');
      window.setTimeout(() => {{
        if (window.Plotly) {{
          document.querySelectorAll(`#${{button.dataset.tab}} .js-plotly-plot`).forEach((plot) => Plotly.Plots.resize(plot));
        }}
      }}, 0);
    }});
  }});
</script>
</body>
</html>"""
    html_path = Path(html_path)
    html_path.parent.mkdir(parents=True, exist_ok=True)
    html_path.write_text(document, encoding="utf-8")
    return html_path.resolve()


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("settings_dir", type=Path)
    parser.add_argument("strategy_file")
    parser.add_argument(
        "weights_excel",
        type=Path,
        help="Workbook containing curr_weight and new_weight",
    )
    parser.add_argument("output_excel", type=Path)
    parser.add_argument(
        "--output-html",
        type=Path,
        help="HTML report path (default: output_excel with an .html suffix)",
    )
    parser.add_argument("--start-date", type=dt.date.fromisoformat, required=True)
    parser.add_argument("--end-date", type=dt.date.fromisoformat, required=True)
    parser.add_argument("--as-of", type=dt.date.fromisoformat)
    parser.add_argument(
        "--scaler",
        type=float,
        help="Explicit strategy scaler override (default: JSON config.pos_scaler)",
    )
    parser.add_argument("--cost-mode", choices=("netted", "sleeve"), default="netted")
    parser.add_argument("--cost-multiplier", type=float, default=1.0)
    parser.add_argument("--holding-lag", type=int, default=2)
    parser.add_argument(
        "--pnl-tenors",
        nargs="+",
        default=list(DEFAULT_PNL_TENORS),
        help="Trailing btmetrics windows, for example: 3m 6m 1y 2y 3y",
    )
    parser.add_argument(
        "--signal-source",
        choices=("generated", "factor-db"),
        default="generated",
        help=(
            "Signal history source. 'generated' loads historical spot/futures "
            "and evaluates signal_store recipes (default); 'factor-db' is the "
            "legacy precomputed-factor adapter."
        ),
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    template = load_strategy_scenario(
        args.settings_dir,
        args.strategy_file,
        scaler=args.scaler,
        scenario_name="template",
    )
    baseline, proposed = load_excel_weight_scenarios(
        template,
        args.weights_excel,
    )
    provider_builder = (
        build_generated_historical_provider
        if args.signal_source == "generated"
        else build_production_factor_provider
    )
    provider = provider_builder(
        [baseline, proposed],
        start_date=args.start_date,
        end_date=args.end_date,
        as_of=args.as_of,
        holding_lag=args.holding_lag,
    )
    comparison = run_strategy_comparison(
        baseline,
        proposed,
        provider,
        cost_mode=args.cost_mode,
        cost_multiplier=args.cost_multiplier,
        pnl_tenors=args.pnl_tenors,
    )
    comparison = trim_comparison_start(comparison, args.start_date)
    output = write_comparison_excel(comparison, args.output_excel)
    html_output = write_comparison_html(
        comparison,
        args.output_html or args.output_excel.with_suffix(".html"),
    )
    print(f"Wrote strategy comparison to {output}")
    print(f"Wrote portfolio HTML report to {html_output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
