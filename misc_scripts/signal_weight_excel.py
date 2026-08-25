"""Round-trip strategy signal weights between JSON settings and Excel.

The workbook schema is intentionally small and stable:

    strategy | factor_name | signal_name | type | curr_weight | new_weight

``strategy`` is the JSON file name (for example ``PTSIM1_AUSPD.json``), and
``factor_name`` is the key inside ``config.factor_repo``.
"""

from __future__ import annotations

import argparse
import json
import math
import os
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, Iterable, Mapping, Sequence

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo


DEFAULT_SHEET_NAME = "signal_weights"
WORKBOOK_COLUMNS = (
    "strategy",
    "factor_name",
    "signal_name",
    "type",
    "curr_weight",
    "new_weight",
)
NEW_FACTOR_DEFAULTS: Mapping[str, Any] = {
    "exec_assets": [],
    "threshold": 0.0,
    "rebal": 1,
    "param": [0.0, 0.0],
}


@dataclass(frozen=True)
class ImportResult:
    """Summary returned by :func:`import_signal_weights_from_excel`."""

    processed_rows: int
    updated_factors: int
    added_factors: int
    skipped_missing_strategies: int
    written_files: tuple[Path, ...]


def _load_strategy_json(path: Path) -> Dict[str, Any]:
    with path.open("r", encoding="utf-8") as handle:
        data = json.load(handle)
    if not isinstance(data, dict):
        raise ValueError(f"Strategy JSON must contain an object: {path}")
    return data


def _factor_repo(data: Mapping[str, Any], source: Path) -> Dict[str, Any]:
    config = data.get("config")
    if not isinstance(config, dict):
        raise ValueError(f"Missing object 'config' in {source}")
    repo = config.get("factor_repo")
    if not isinstance(repo, dict):
        raise ValueError(f"Missing object 'config.factor_repo' in {source}")
    return repo


def _strategy_files(settings_dir: Path) -> Iterable[Path]:
    return sorted(settings_dir.glob("*.json"), key=lambda path: path.name.lower())


def export_signal_weights_to_excel(
    settings_dir: str | Path,
    excel_path: str | Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> Path:
    """Export all ``config.factor_repo`` entries to a formatted Excel table.

    The source JSON files are only read. The returned path is absolute.
    """

    settings_dir = Path(settings_dir)
    excel_path = Path(excel_path)
    if not settings_dir.is_dir():
        raise FileNotFoundError(f"Settings directory does not exist: {settings_dir}")

    rows: list[tuple[Any, ...]] = []
    for strategy_path in _strategy_files(settings_dir):
        data = _load_strategy_json(strategy_path)
        repo = _factor_repo(data, strategy_path)
        for factor_name, signal in repo.items():
            if not isinstance(signal, dict):
                raise ValueError(
                    f"Factor '{factor_name}' must contain an object in {strategy_path}"
                )
            missing = [key for key in ("name", "type", "weight") if key not in signal]
            if missing:
                raise ValueError(
                    f"Factor '{factor_name}' in {strategy_path} is missing: "
                    + ", ".join(missing)
                )
            rows.append(
                (
                    strategy_path.name,
                    str(factor_name),
                    signal["name"],
                    signal["type"],
                    signal["weight"],
                    signal["weight"],
                )
            )

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = sheet_name
    worksheet.append(WORKBOOK_COLUMNS)
    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")

    worksheet.freeze_panes = "A2"
    worksheet.sheet_view.showGridLines = False
    worksheet.column_dimensions["A"].width = 30
    worksheet.column_dimensions["B"].width = 32
    worksheet.column_dimensions["C"].width = 30
    worksheet.column_dimensions["D"].width = 18
    worksheet.column_dimensions["E"].width = 15
    worksheet.column_dimensions["F"].width = 15
    for column in ("E", "F"):
        for cell in worksheet[column][1:]:
            cell.number_format = "0.###############"
    # ``new_weight`` is the user-controlled proposal input.
    for cell in worksheet["F"][1:]:
        cell.font = Font(color="0000FF")

    if rows:
        table = Table(displayName="SignalWeights", ref=f"A1:F{len(rows) + 1}")
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)
    else:
        worksheet.auto_filter.ref = "A1:F1"

    excel_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(excel_path)
    workbook.close()
    return excel_path.resolve()


def _normalized_headers(worksheet: Any) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for column, cell in enumerate(worksheet[1], start=1):
        if cell.value is None:
            continue
        name = str(cell.value).strip().lower().replace(" ", "_")
        if name in headers:
            raise ValueError(f"Duplicate Excel column: {name}")
        headers[name] = column
    missing = [column for column in WORKBOOK_COLUMNS if column not in headers]
    if missing:
        raise ValueError("Excel sheet is missing columns: " + ", ".join(missing))
    return headers


def _required_text(value: Any, column: str, row_number: int) -> str:
    if value is None or not str(value).strip():
        raise ValueError(f"Row {row_number}: '{column}' must not be blank")
    return str(value).strip()


def _weight(value: Any, row_number: int, column: str) -> float:
    if isinstance(value, bool):
        raise ValueError(f"Row {row_number}: '{column}' must be numeric")
    try:
        result = float(value)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"Row {row_number}: '{column}' must be numeric") from exc
    if not math.isfinite(result):
        raise ValueError(f"Row {row_number}: '{column}' must be finite")
    return result


def _strategy_path(settings_dir: Path, value: Any, row_number: int) -> Path:
    strategy = _required_text(value, "strategy", row_number)
    if Path(strategy).name != strategy or strategy in {".", ".."}:
        raise ValueError(f"Row {row_number}: 'strategy' must be a file name")
    if Path(strategy).suffix.lower() != ".json":
        strategy += ".json"
    return settings_dir / strategy


def _atomic_write_json(path: Path, data: Mapping[str, Any]) -> None:
    temp_name: str | None = None
    try:
        with tempfile.NamedTemporaryFile(
            mode="w",
            encoding="utf-8",
            newline="\n",
            dir=path.parent,
            prefix=f".{path.name}.",
            suffix=".tmp",
            delete=False,
        ) as handle:
            temp_name = handle.name
            json.dump(data, handle, ensure_ascii=False, indent=4)
            handle.write("\n")
        os.replace(temp_name, path)
    except Exception:
        if temp_name is not None:
            Path(temp_name).unlink(missing_ok=True)
        raise


def generate_strategy_json_from_excel(
    strategy_json: str | Path,
    excel_path: str | Path,
    output_json: str | Path | None = None,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> Path:
    """Create a proposed strategy JSON from one workbook strategy.

    The source JSON is read-only. Matching workbook rows replace its
    ``config.factor_repo`` so the generated JSON represents the same proposed
    signal set used by the scenario backtest. Existing factors preserve all
    fields other than ``name``, ``type`` and ``weight``; new factors receive
    :data:`NEW_FACTOR_DEFAULTS`. JSON ``weight`` comes from ``new_weight``.

    When ``output_json`` is omitted, the file is written beside the workbook
    as ``<source_stem>_proposed.json``.
    """

    strategy_path = Path(strategy_json)
    excel_path = Path(excel_path)
    if not strategy_path.is_file():
        raise FileNotFoundError(f"Strategy JSON does not exist: {strategy_path}")
    if strategy_path.suffix.lower() != ".json":
        raise ValueError(f"Strategy file must be JSON: {strategy_path}")
    if output_json is None:
        output_path = excel_path.parent / f"{strategy_path.stem}_proposed.json"
    else:
        output_path = Path(output_json)
    if output_path.suffix.lower() != ".json":
        raise ValueError(f"Output file must be JSON: {output_path}")
    if output_path.resolve() == strategy_path.resolve():
        raise ValueError("output_json must not overwrite the source strategy JSON")

    data = _load_strategy_json(strategy_path)
    source_repo = _factor_repo(data, strategy_path)
    accepted_names = {strategy_path.name.lower(), strategy_path.stem.lower()}

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Excel sheet '{sheet_name}' not found; available sheets: "
                + ", ".join(workbook.sheetnames)
            )
        worksheet = workbook[sheet_name]
        headers = _normalized_headers(worksheet)
        proposed_repo: Dict[str, Any] = {}

        for row_number in range(2, worksheet.max_row + 1):
            strategy_value = worksheet.cell(
                row_number, headers["strategy"]
            ).value
            if strategy_value is None:
                continue
            if str(strategy_value).strip().lower() not in accepted_names:
                continue

            factor_name = _required_text(
                worksheet.cell(row_number, headers["factor_name"]).value,
                "factor_name",
                row_number,
            )
            if factor_name in proposed_repo:
                raise ValueError(
                    f"Row {row_number}: duplicate factor '{factor_name}' for "
                    f"{strategy_path.name}"
                )
            signal_name = _required_text(
                worksheet.cell(row_number, headers["signal_name"]).value,
                "signal_name",
                row_number,
            )
            signal_type = _required_text(
                worksheet.cell(row_number, headers["type"]).value,
                "type",
                row_number,
            )
            signal_weight = _weight(
                worksheet.cell(row_number, headers["new_weight"]).value,
                row_number,
                "new_weight",
            )

            if factor_name in source_repo:
                source_signal = source_repo[factor_name]
                if not isinstance(source_signal, dict):
                    raise ValueError(
                        f"Factor '{factor_name}' must contain an object in "
                        f"{strategy_path}"
                    )
                signal = dict(source_signal)
                signal["name"] = signal_name
                signal["type"] = signal_type
                signal["weight"] = signal_weight
            else:
                signal = {
                    "name": signal_name,
                    "type": signal_type,
                    "exec_assets": list(NEW_FACTOR_DEFAULTS["exec_assets"]),
                    "threshold": NEW_FACTOR_DEFAULTS["threshold"],
                    "rebal": NEW_FACTOR_DEFAULTS["rebal"],
                    "param": list(NEW_FACTOR_DEFAULTS["param"]),
                    "weight": signal_weight,
                }
            proposed_repo[factor_name] = signal
    finally:
        workbook.close()

    if not proposed_repo:
        raise ValueError(
            f"No rows found for strategy '{strategy_path.name}' in {excel_path}"
        )
    data["config"]["factor_repo"] = proposed_repo
    output_path.parent.mkdir(parents=True, exist_ok=True)
    _atomic_write_json(output_path, data)
    return output_path.resolve()


def import_signal_weights_from_excel(
    settings_dir: str | Path,
    excel_path: str | Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
) -> ImportResult:
    """Apply Excel signal rows to existing strategy JSON files.

    Existing factors keep every field except ``name``, ``type``, and ``weight``;
    JSON ``weight`` is populated from workbook ``new_weight``. ``curr_weight``
    is retained for backtest comparison and is not written to JSON.
    New factors receive ``exec_assets=[]``, ``threshold=0.0``, ``rebal=1``, and
    ``param=[0.0, 0.0]``. A row whose strategy file does not exist is skipped.

    All rows are validated before any JSON file is written. Duplicate
    ``(strategy, factor_name)`` rows are rejected to prevent ambiguous updates.
    """

    excel_path = Path(excel_path)
    settings_dir = Path(settings_dir)
    if not settings_dir.is_dir():
        raise FileNotFoundError(f"Settings directory does not exist: {settings_dir}")

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"Excel sheet '{sheet_name}' not found; available sheets: "
                + ", ".join(workbook.sheetnames)
            )
        worksheet = workbook[sheet_name]
        headers = _normalized_headers(worksheet)

        updates: Dict[Path, list[tuple[str, str, str, float]]] = {}
        seen: set[tuple[str, str]] = set()
        processed_rows = 0
        skipped = 0

        for row_number in range(2, worksheet.max_row + 1):
            values = {
                column: worksheet.cell(row_number, index).value
                for column, index in headers.items()
            }
            if all(values[column] is None for column in WORKBOOK_COLUMNS):
                continue
            processed_rows += 1
            strategy_path = _strategy_path(
                settings_dir, values["strategy"], row_number
            )
            if not strategy_path.is_file():
                skipped += 1
                continue

            factor_name = _required_text(
                values["factor_name"], "factor_name", row_number
            )
            signal_name = _required_text(
                values["signal_name"], "signal_name", row_number
            )
            signal_type = _required_text(values["type"], "type", row_number)
            signal_weight = _weight(values["new_weight"], row_number, "new_weight")

            key = (strategy_path.name.lower(), factor_name)
            if key in seen:
                raise ValueError(
                    f"Row {row_number}: duplicate strategy/factor pair "
                    f"'{strategy_path.name}' / '{factor_name}'"
                )
            seen.add(key)
            updates.setdefault(strategy_path, []).append(
                (factor_name, signal_name, signal_type, signal_weight)
            )
    finally:
        workbook.close()

    loaded: Dict[Path, Dict[str, Any]] = {}
    updated_factors = 0
    added_factors = 0
    for strategy_path, factor_updates in updates.items():
        data = _load_strategy_json(strategy_path)
        repo = _factor_repo(data, strategy_path)
        for factor_name, signal_name, signal_type, signal_weight in factor_updates:
            if factor_name in repo:
                signal = repo[factor_name]
                if not isinstance(signal, dict):
                    raise ValueError(
                        f"Factor '{factor_name}' must contain an object in "
                        f"{strategy_path}"
                    )
                signal["name"] = signal_name
                signal["type"] = signal_type
                signal["weight"] = signal_weight
                updated_factors += 1
            else:
                repo[factor_name] = {
                    "name": signal_name,
                    "type": signal_type,
                    "exec_assets": list(NEW_FACTOR_DEFAULTS["exec_assets"]),
                    "threshold": NEW_FACTOR_DEFAULTS["threshold"],
                    "rebal": NEW_FACTOR_DEFAULTS["rebal"],
                    "param": list(NEW_FACTOR_DEFAULTS["param"]),
                    "weight": signal_weight,
                }
                added_factors += 1
        loaded[strategy_path] = data

    for strategy_path, data in loaded.items():
        _atomic_write_json(strategy_path, data)

    return ImportResult(
        processed_rows=processed_rows,
        updated_factors=updated_factors,
        added_factors=added_factors,
        skipped_missing_strategies=skipped,
        written_files=tuple(path.resolve() for path in sorted(loaded)),
    )


def _build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    subparsers = parser.add_subparsers(dest="command", required=True)

    export_parser = subparsers.add_parser("export", help="JSON settings to Excel")
    export_parser.add_argument("settings_dir", type=Path)
    export_parser.add_argument("excel_path", type=Path)
    export_parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME)

    import_parser = subparsers.add_parser("import", help="Excel rows to JSON settings")
    import_parser.add_argument("settings_dir", type=Path)
    import_parser.add_argument("excel_path", type=Path)
    import_parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME)

    generate_parser = subparsers.add_parser(
        "generate",
        help="Create one proposed JSON without changing its source JSON",
    )
    generate_parser.add_argument("strategy_json", type=Path)
    generate_parser.add_argument("excel_path", type=Path)
    generate_parser.add_argument(
        "output_json",
        type=Path,
        nargs="?",
        help="Default: <Excel folder>/<strategy stem>_proposed.json",
    )
    generate_parser.add_argument("--sheet-name", default=DEFAULT_SHEET_NAME)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _build_parser().parse_args(argv)
    if args.command == "export":
        output = export_signal_weights_to_excel(
            settings_dir=args.settings_dir,
            excel_path=args.excel_path,
            sheet_name=args.sheet_name,
        )
        print(f"Exported signal weights to {output}")
        return 0

    if args.command == "generate":
        output = generate_strategy_json_from_excel(
            strategy_json=args.strategy_json,
            excel_path=args.excel_path,
            output_json=args.output_json,
            sheet_name=args.sheet_name,
        )
        print(f"Generated proposed strategy JSON at {output}")
        return 0

    result = import_signal_weights_from_excel(
        settings_dir=args.settings_dir,
        excel_path=args.excel_path,
        sheet_name=args.sheet_name,
    )
    print(
        f"Processed {result.processed_rows} rows: "
        f"updated={result.updated_factors}, added={result.added_factors}, "
        f"skipped_missing_strategy={result.skipped_missing_strategies}, "
        f"written_files={len(result.written_files)}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
